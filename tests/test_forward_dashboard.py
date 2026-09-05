from datetime import date, timedelta
import sqlite3
from pathlib import Path

from streamlit.testing.v1 import AppTest

from range_finder.forward_test.results import load_results
from range_finder.forward_test.runner import run_study
from range_finder.forward_test.store import Store
from range_finder.tests.forward_fixtures import Clock, FixtureProvider
from range_finder.trading_week import trading_week


def test_dashboard_filters_export_and_real_app_route(tmp_path,monkeypatch):
    week=trading_week(date(2026,9,7)); clock=Clock(week.capture_start)
    store=Store(sqlite3.connect(tmp_path/'view.sqlite',isolation_level=None))
    store.migrate(); store.register('fixture',str(week.monday),clock(),{'fixture':True})
    provider=FixtureProvider(clock)
    run_study(store,provider,'fixture',clock=clock,model_version='fixture-v1')
    clock.value=week.evaluation_close+timedelta(hours=3)
    run_study(store,provider,'fixture',clock=clock,model_version='fixture-v1')
    rows=load_results(store,'fixture')
    runs=store.query('SELECT * FROM ft_runs ORDER BY started_at DESC')
    store.close()
    import ui_forward_test
    monkeypatch.setattr(ui_forward_test,'load_snapshot',lambda:(rows,runs,[]))
    at=AppTest.from_string('from ui_forward_test import render_forward_test\nrender_forward_test()').run(timeout=60)
    assert not at.exception
    assert at.metric[0].value=='100.0%'
    assert len(at.dataframe[1].value)==64
    at.multiselect(key='ft_filter_ticker').select('AMD').run()
    assert not at.exception and len(at.dataframe[1].value)==16
    at.multiselect(key='ft_filter_model').select('M2_vix').run()
    assert len(at.dataframe[1].value)==4
    assert len(at.get('download_button'))==1
    # The real app routes this view before its credential/GEX market-data gate.
    app=AppTest.from_file(str(Path(__file__).resolve().parents[1]/'streamlit_app.py'))
    app.query_params['tab']='forward'
    app.run(timeout=60)
    assert not app.exception
    assert any(t.value=='Forward Test' for t in app.title)


def test_unactivated_dashboard_is_read_only(monkeypatch):
    import ui_forward_test
    def missing():raise RuntimeError('fixture missing schema')
    monkeypatch.setattr(ui_forward_test,'load_snapshot',missing)
    at=AppTest.from_string('from ui_forward_test import render_forward_test\nrender_forward_test()').run()
    assert not at.exception
    assert 'not available' in at.info[0].value


def test_registered_empty_study_filters_and_excel(monkeypatch):
    import ui_forward_test
    from io import BytesIO
    from openpyxl import load_workbook
    from range_finder.forward_test.results import build_workbook
    studies=[{'study_id':'prospective-only','start_week':'2026-09-07'}]
    monkeypatch.setattr(ui_forward_test,'load_snapshot',lambda:([],[],studies))
    at=AppTest.from_string('from ui_forward_test import render_forward_test\nrender_forward_test()').run()
    assert not at.exception
    assert 'No frozen predictions' in at.info[0].value
    assert any('2026-09-07' in c.value for c in at.caption)
    assert at.multiselect(key='ft_filter_study_id').options==['prospective-only']
    at.multiselect(key='ft_filter_ticker').select('SPX').run()
    assert not at.exception and at.metric[0].value=='—'
    assert len(at.get('download_button'))==1
    wb=load_workbook(BytesIO(build_workbook([])))
    assert wb.sheetnames==['Scoreboard','Breach details','Read me']
    assert wb['Scoreboard'].max_row==4  # headers only; no synthetic records
