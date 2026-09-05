"""
Spread Finder tab UI — weekly credit spread planning with forecast,
GEX context, and interactive strike maps.
Extracted from streamlit_app.py.
"""
from __future__ import annotations

from datetime import date as date_cls, datetime, timedelta, timezone

import numpy as np
import pandas as pd
import streamlit as st

from theme import COLORS
from models import GEXData
from ui_history import _is_weekly_freeze_day

from range_finder.gex_bridge import (
    GEXContext, extract_gex_context, save_gex_to_range_finder,
    adjust_spread_with_gex, regime_to_gex_flag, reconcile_gex_warnings,
)
from range_finder.data_collector import (
    fetch_spx_vix as rf_fetch_spx_vix, save_spx_vix as rf_save_spx_vix,
    fetch_fred_macro as rf_fetch_fred_macro, save_fred_macro as rf_save_fred_macro,
    build_event_flags as rf_build_event_flags,
    get_weekly_spx as rf_get_weekly_spx,
    fred_key_status as rf_fred_key_status,
    capture_and_save_monday_anchor as rf_capture_monday_anchor,
    FRED_API_KEY as RF_FRED_API_KEY,
)
from range_finder.feature_builder import (
    build_features as rf_build_features,
    get_features as rf_get_features,
    assess_path_provenance as rf_assess_path_provenance,
)
from range_finder.gex_policy import (
    GEX_LIVE_SPREAD_INFLUENCE_ENABLED,
    GEX_NORMALIZED_FEATURE,
    live_spread_feature_columns,
    uses_disabled_gex_feature,
)
from range_finder.har_model import (
    MODEL_SPECS as RF_MODEL_SPECS, PI_ALPHA as RF_PI_ALPHA,
    GEX_MIN_WEEKS_FOR_FIT as RF_GEX_MIN_WEEKS,
    feature_has_enough_data as rf_feature_has_enough_data,
    production_feature_columns,
    fit_validation_and_production as rf_fit_validation_and_production,
    forecast_next_week as rf_forecast_next_week,
    estimate_side_share_quantile as rf_estimate_side_share_quantile,
    save_model as rf_save_model, load_model as rf_load_model,
)
from range_finder.spread_levels import (
    build_spread_plan as rf_build_spread_plan,
    build_spread_tiers as rf_build_spread_tiers,
    update_outcome as rf_update_outcome,
    MIN_CREDIT_RATIO,
    TICKER_CONFIG as RF_TICKER_CONFIG,
    SpreadPlan,
    SpreadTier,
)


# ─────────────────────────────────────────────────────────────────────────────
# Spread Finder Tab — HAR placement with GEX retained as research context
# ─────────────────────────────────────────────────────────────────────────────

from range_finder.recommendations import (
    build_recommendations, displayed_tier, tier_bands, chain_entry_to_quotes,
)

from theme import SF_BG, SF_BULL, SF_BEAR, SF_NEUT, SF_WARN, SF_CARD


@st.cache_resource(ttl=3600)
def _get_rf_conn():
    """Get or create the range finder Postgres connection.

    1-hour TTL lets a long-idle Streamlit Cloud process drop its backend
    connection so Neon can auto-suspend the compute. Without a TTL the
    cached connection persists for the full process lifetime, holding
    the endpoint warm even when the user has no tab open. ``init_all_tables``
    is idempotent (CREATE TABLE IF NOT EXISTS + ALTER ... IF NOT EXISTS),
    so re-running it on each reconnect is safe and fast (<100 ms).
    """
    from range_finder.db import get_connection, init_all_tables
    conn = get_connection()
    init_all_tables(conn)
    return conn


@st.cache_data(ttl=600, show_spinner=False)
def _cached_pi_coverage(ticker: str) -> dict:
    """Empirical PI-coverage summary from scored spread_log rows.

    One small aggregate SELECT, cached 10 min — the underlying data only
    changes once a week (Monday cron logs the new plan and scores last
    week's), so reruns cost nothing. See range_finder/calibration.py.
    """
    from range_finder.calibration import weekly_pi_coverage
    return weekly_pi_coverage(_get_rf_conn(), ticker=ticker)


@st.cache_data(ttl=3600, show_spinner=False)
def _cached_side_share_q(ticker: str) -> "float | None":
    """Empirical side-share quantile for per-side band placement, cached 1h
    (the weekly OHLC it derives from changes once a week). None → callers
    keep the legacy symmetric /2 split, so a DB hiccup degrades gracefully
    to pre-side-share behavior instead of erroring the tab."""
    try:
        from range_finder.feature_builder import _load_weekly_for_ticker
        est = rf_estimate_side_share_quantile(
            _load_weekly_for_ticker(_get_rf_conn(), ticker))
        return est["q"]
    except Exception:
        return None


@st.cache_data(ttl=300, show_spinner=False)
def _live_vol_close(proxy: str) -> float | None:
    """Latest vol-proxy (VIX/VXN) close, refreshed every 5 minutes.

    Returns None on an empty/failed fetch — the caller must treat None as
    "no live VIX" and NOT fabricate a number, because this value drives the
    VIX regime-shift circuit breaker: a stale or invented VIX silently
    disables the breaker exactly when a live spike should trip it. A 5-min
    TTL (vs the old cache-once-per-session behavior) means a mid-session
    spike is actually seen while still sparing yfinance on every rerun.
    """
    try:
        vp_hist = yf.Ticker(proxy).history(period="5d")
        if not vp_hist.empty:
            return round(float(vp_hist["Close"].dropna().iloc[-1]), 2)
    except Exception:
        pass
    return None


@st.cache_data(ttl=600, show_spinner=False)
def _cached_rf_get_features(_conn, ticker: str = "SPX"):
    """model_features table for one ticker, cached for 10 minutes.

    Streamlit reruns the Spread Finder render every time any widget changes
    — dropdown, number input, slider, risk-tier button — and the raw
    ``rf_get_features`` helper runs a full ``SELECT *`` on ``model_features``
    each call. On Neon's free tier that single query dominated the
    CU-hour bill. The 10 min TTL is well under the weekly cadence at which
    features actually change; the Refresh/Rebuild/Weekly Setup paths call
    ``.clear()`` below to force a reload the moment new data lands.

    The ``_conn`` underscore tells Streamlit to skip hashing the connection
    object (psycopg2 connections aren't hashable).

    The window is pinned to har_model.TRAIN_WINDOW_YEARS (same as the cron
    fit path) so deeper weekly_spx backfills — e.g. the 10y history
    experiment — can't silently change what the UI trains or forecasts on.
    """
    from range_finder.har_model import train_window_min_date
    return rf_get_features(_conn, min_date=train_window_min_date(),
                           exclude_covid=True, ticker=ticker)


@st.cache_resource(ttl=3600, show_spinner=False)
def _cached_rf_load_model(model_name: str, ticker: str):
    """Load a HAR fit from saved_models, cached for 1 hour across sessions.

    The session-state guards in the render path stop the unpickle from
    re-running within a single session, but every fresh session was
    re-fetching the multi-MB BYTEA blob from Neon — up to 8 SELECTs
    (4 specs × 2 tickers) per session lifetime. This wrapper caches the
    unpickled payload at the Streamlit-instance level so users sharing
    the same instance share the load.

    Uses ``@st.cache_resource`` (not ``cache_data``) because the unpickled
    statsmodels result wrapper is a non-trivial Python object that
    shouldn't be re-deserialized per call. The cache key is
    (model_name, ticker), so SPX and XSP fits stay independent — same as
    the underlying ``saved_models`` PK.

    Save paths in this module call ``_cached_rf_load_model.clear()``
    after a successful refit so the new fit is picked up immediately
    instead of waiting for the TTL to expire. The 1-hour TTL bridges
    the case where ``scheduled_snapshot.py`` (Mondays 9:30 AM ET) writes
    a new fit without any UI interaction.
    """
    from range_finder.db import get_connection
    # Open a short-lived connection rather than reusing the cached
    # Streamlit connection — _cached_rf_load_model is keyed on
    # (model_name, ticker), so passing in the cached `conn` would couple
    # cache invalidation to connection identity. The blob is pulled once
    # per (spec, ticker) per hour, so the extra connect cost is trivial.
    conn = get_connection()
    try:
        return rf_load_model(model_name, conn=conn, ticker=ticker)
    finally:
        try:
            conn.close()
        except Exception:
            pass


@st.cache_data(ttl=900, show_spinner=False)
def _cached_weekly_setup(_conn, week_start: str, ticker: str):
    """Look up the weekly_setup row for (week_start, ticker), cached 15 min.

    Replaces a hand-rolled session-state miss-cache that still hit Neon
    on every fresh session. Cross-session caching means the SELECT only
    fires once per Streamlit instance per 15-minute window, regardless
    of how many users / tabs / reruns happen.

    Returns (monday_open, monday_vix) on hit, or None on miss. Both hits
    AND misses are cached — the bespoke logic this replaced cached only
    misses, so hits still re-queried on every fresh session.

    Save paths (``do_weekly`` / ``do_save_gex``) call ``.clear()`` to
    invalidate, mirroring the existing ``_cached_rf_get_features.clear()``
    pattern.
    """
    cur = _conn.cursor()
    cur.execute(
        "SELECT monday_open, monday_vix FROM weekly_setup WHERE week_start = ? AND ticker = ?",
        (week_start, ticker),
    )
    row = cur.fetchone()
    if row and row[0]:
        return (row[0], row[1])
    return None


def _spread_finder_target_friday(ref_date: "date_cls | None" = None) -> "date_cls":
    """Return the calendar Friday of the week the Spread Finder is planning for.

    On Mon-Thu we're inside a live trading week — traders entering new
    credit spreads want *this* week's Friday (the one that's 0-4 days
    away).  On Fri-Sun the current week is effectively done, so we roll
    forward to next Monday's week and pick its Friday.  The same rule is
    applied in ``_render_spread_finder_tab`` when deriving ``week_start``
    so both stay in sync.
    """
    today = ref_date or date_cls.today()
    wd = today.weekday()
    if wd <= 3:  # Mon-Thu → this week's Monday
        monday = today - timedelta(days=wd)
    else:        # Fri-Sun → next Monday
        monday = today + timedelta(days=(7 - wd))
    return monday + timedelta(days=4)


def find_spread_finder_friday_exp(
    avail: "list[str]",
    ref_date: "date_cls | None" = None,
) -> "str | None":
    """Listed end-of-week expiration for the UI's planned trading week."""
    from range_finder.trading_week import trading_week, listed_week_expiration
    return listed_week_expiration(avail, trading_week(_spread_finder_target_friday(ref_date)))


def _chain_entry_to_quotes(entry: dict) -> dict:
    return chain_entry_to_quotes(entry)


def _build_chain_quotes_for_spreads(
    data: GEXData,
    ticker: str,
    ref_date: "date_cls | None" = None,
) -> tuple[dict, str | None]:
    """Build a strike→{call_bid, call_ask, put_bid, put_ask} lookup from the
    Friday chain that matches the Spread Finder's planned week.

    The target expiration is anchored to *the week the spread finder is
    forecasting* (see ``_spread_finder_target_friday``), not to "whichever
    expiration the user happened to pick in the sidebar".  Before this was
    added, a user who had ``0DTE`` or ``Tomorrow`` selected would see the
    spread finder silently fall back to today's chain — producing $0.00
    credits for far-OTM weekly strikes because it was pricing 0-DTE puts
    instead of Friday weeklies.  The pre-fetch in ``fetch_all_data`` makes
    sure the right chain is always in ``data.chain_cache`` regardless of
    sidebar state, and this function just looks up that exact Friday.

    Returns (quotes_dict, selected_expiration_str_or_None).  When the
    correct Friday isn't available we return empty so ``build_spread_side``
    falls back cleanly to its BSM estimator (the UI caption tells the user
    we're on BSM rather than market quotes).
    """
    if not data.chain_cache:
        return {}, None

    # Resolve the expiration we SHOULD be looking at. Prefer the full
    # expiration universe from data.avail so holiday-shifted Fridays can
    # still match; fall back to whatever's already in the chain cache.
    avail = list(getattr(data, "avail", None) or [])
    if not avail:
        avail = sorted({exp for (t, exp) in data.chain_cache if t == ticker})

    target_exp = find_spread_finder_friday_exp(avail, ref_date=ref_date)
    if target_exp is None:
        return {}, None

    entry = data.chain_cache.get((ticker, target_exp))
    if not entry or entry.get("status") != "ok":
        # The right Friday isn't cached — don't silently substitute another
        # expiration (that's exactly how we used to end up pricing weekly
        # spreads off today's 0DTE chain).  Let the caller fall back to BSM.
        return {}, None

    return _chain_entry_to_quotes(entry), target_exp


def _tradier_token() -> str:
    """Tradier API token from secrets/env (mirrors streamlit_app.get_credentials)."""
    import os
    try:
        tok = st.secrets.get("TRADIER_TOKEN", "")
    except Exception:
        tok = ""
    return tok or os.environ.get("TRADIER_TOKEN", "")


def _export_chain_quotes(ticker: str, ref_date: "date_cls | None" = None) -> tuple[dict, str | None]:
    """Fetch the Spread-Finder-planned-Friday chain for ONE ticker and build the
    strike -> {bid/ask} lookup, so the multi-ticker Excel export snaps to the
    SAME real listed strikes the live tab does.

    Non-active export rows are built from persisted state and don't have the
    active ticker's ``data.chain_cache``, so without this they round to the
    nominal strike_increment — which drifts from the tradeable grid (e.g. AMD
    557.5 on its $2.5 far-OTM grid vs a nominal 556 that isn't even listed).

    Returns ``({}, None)`` on any failure (no token, no Friday listed, fetch
    error) so the caller degrades cleanly to nominal/BSM strikes — exactly
    today's behavior — instead of blocking the export.
    """
    token = _tradier_token()
    if not token:
        return {}, None
    try:
        from streamlit_app import get_expirations_cached
        avail = get_expirations_cached(token, ticker) or []
    except Exception:
        avail = []
    target_exp = find_spread_finder_friday_exp(avail, ref_date=ref_date)
    if not target_exp:
        return {}, None
    try:
        from phase1.data_client import TradierDataClient
        entry = TradierDataClient(token=token).get_chain_once(ticker, target_exp)
    except Exception:
        return {}, None
    if not entry or entry.get("status") != "ok":
        return {}, None
    quotes = _chain_entry_to_quotes(entry)
    return (quotes, target_exp) if quotes else ({}, None)


# ─────────────────────────────────────────────────────────────────────────────
# Weekly forward-test workbook — ALL tickers, one week-named tab
# ─────────────────────────────────────────────────────────────────────────────
# One export carries the always-present default instruments plus any tickers
# the user has added to the export list, one row each, on a sheet named after
# the planning week's Monday. The intended workflow:
#
#   • First download = your MASTER workbook. It ships with a Scoreboard and
#     an empty utility sheet (_FT_BLANK_SHEET) the Scoreboard needs.
#   • Every following week: download the new export, right-click its week
#     tab → Move or Copy → into the master (any position).
#   • The Scoreboard is TICKER-KEYED, not positional: it matches every week
#     tab's rows by Instrument name (SUMIF/COUNTIFS over INDIRECT built from
#     the pre-listed Monday tab names, existence-sanitized in its hidden
#     column L — see _write_ft_scoreboard). The ticker mix and row order are
#     free to differ week to week; each instrument's score simply accumulates
#     over the weeks it appears. Typing a ticker into a blank Scoreboard slot
#     starts tracking it retroactively across every tab already in the book.
#   • Upgrading an old positional (3D-sum) master: delete its Scoreboard and
#     the WeeksStart/WeeksEnd bookends (obsolete), then Move/Copy a fresh
#     export's Scoreboard AND FTBlank sheets in (Ctrl+click both tabs) —
#     they're self-contained as a pair. Re-type any previously tracked
#     tickers into blank Instrument slots; they re-score retroactively.
#     Caveat: pre-rework week tabs carry the old unguarded Scored?/flag
#     formulas, so a band-less row whose close was filled shows a 0 flag
#     there — the Scoreboard's denominators filter those out (they require
#     the tier's Low band cell to hold a number), but old tabs' Weeks
#     Scored can still count band-less rows. Re-paste a new tab's O..S
#     formulas over an old tab's rows to fully repair it.
#
# Per-week sheet layout:
#   A Week(Mon) · B Instrument · C Class · D Ref/Open · E Weekly Close
#   (user fills after Friday) · F Prev Close · G..N four tier bands
#   (Low/High) · O..R CLOSE INSIDE? per tier (formulas) · S Scored?
#   (formula) · T Notes

# Always-present defaults, in display order: the index/ETF defaults followed by
# the single-name defaults the Monday cron fits (NVDA/JPM/CAT). Mirrors
# QUICK_TICKERS in ui_theme — keep the two in sync. The active/searched ticker
# and any the user has added (session-state list under _FT_XLSX_EXTRA_KEY) are
# appended.
_FT_DEFAULT_TICKERS = ["SPX", "XSP", "SPY", "QQQ", "NDX", "NVDA", "JPM", "CAT"]
# Class-label overrides for the workbook "Class" column; anything else falls
# back to the ticker_config category (Index/ETF/Stock).
_FT_CLASS = {"SPX": "Index", "XSP": "Index", "SPY": "ETF",
             "QQQ": "ETF", "NDX": "Index"}
_FT_FIRST_DATA_ROW = 6          # first instrument row on each week sheet
_FT_MAX_TICKERS = 40            # rows the Scoreboard scans on each week tab (B6:B45)
_FT_LAST_DATA_ROW = _FT_FIRST_DATA_ROW + _FT_MAX_TICKERS - 1
_FT_SB_FIRST_ROW = 6            # first instrument slot on the Scoreboard
_FT_SB_SLOTS = 40               # slots incl. blanks the user can type a ticker into
_FT_SB_LAST_ROW = _FT_SB_FIRST_ROW + _FT_SB_SLOTS - 1
_FT_SB_ALL_ROW = _FT_SB_LAST_ROW + 1
_FT_SB_WEEKS_BACK = 260         # Monday tab names pre-listed behind the export week
                                # (5y — a fresh Scoreboard must cover every tab an
                                # older master could have accumulated)
_FT_SB_WEEKS_FWD = 156          # ...and ahead of it: the 3y runway before the
                                # Scoreboard (which warns as it nears) needs swapping
                                # for a fresh export's
_FT_BLANK_SHEET = "FTBlank"     # empty utility sheet not-yet-pasted weeks resolve to
_FT_XLSX_EXTRA_KEY = "_sf_xlsx_extra"   # session-state list of user-added tickers


def _ft_class(ticker: str) -> str:
    """Asset-class label for the workbook 'Class' column (Index/ETF/Stock)."""
    t = (ticker or "").upper()
    if t in _FT_CLASS:
        return _FT_CLASS[t]
    try:
        from phase1.ticker_config import get_config
        return str(get_config(t)["category"]).title()
    except Exception:
        return "Stock"


def _default_model_for_ticker(ticker: str) -> str:
    """Per-ticker default model spec.

    Walk-forward OOS validation showed M3_extended is the strongest forecaster
    on index products (SPX/XSP/QQQ/NDX/SPY...), while the simpler M2_vix wins on
    single names (AMZN/AMD/...), whose weekly range isn't well explained by the
    market-wide event-count / S&P return-lag features that M3 adds. So index/ETF
    products default to M3_extended and single stocks to M2_vix. The user can
    still override via the dropdown — the choice is remembered per ticker.
    """
    return "M2_vix" if _ft_class(ticker) == "Stock" else "M3_extended"


import re as _re

# A valid exchange/OCC ticker root: 1-6 chars, A-Z 0-9 and a dot only. This
# is the primary defense against spreadsheet formula injection — the export
# tickers can arrive via URL state (ui_url_state), so a value like
# ``=cmd|'/c calc'!A1`` or ``@SUM(...)`` must never reach a cell. Anything
# outside this charset is rejected at the door.
_TICKER_RE = _re.compile(r"^[A-Z0-9.]{1,6}$")


def _valid_ticker_symbol(ticker: str) -> bool:
    return bool(_TICKER_RE.match((ticker or "").upper()))


def _xlsx_safe_cell(text) -> str:
    """Neutralize a leading spreadsheet formula trigger on any string cell.

    Defense in depth behind _valid_ticker_symbol: Excel/Sheets treat a cell
    beginning with = + - @ (or tab/CR) as a formula. Prefix such a value with
    an apostrophe so it renders literally. Non-strings pass through unchanged.
    """
    if not isinstance(text, str) or not text:
        return text
    return "'" + text if text[0] in ("=", "+", "-", "@", "\t", "\r") else text


def _xlsx_extra_list() -> list:
    """User-added export tickers (session-state; deduped, defaults excluded).

    Filtered through _valid_ticker_symbol on the way out so a tampered
    session/URL value can never reach the workbook builder even if it was
    persisted before validation tightened."""
    raw = st.session_state.setdefault(_FT_XLSX_EXTRA_KEY, [])
    return [t for t in raw if _valid_ticker_symbol(t)]


def _xlsx_add_extra(ticker: str) -> None:
    """Append a ticker to the export list (no-op for defaults / duplicates /
    anything that isn't a valid exchange symbol)."""
    t = (ticker or "").upper()
    if not t or t in _FT_DEFAULT_TICKERS or not _valid_ticker_symbol(t):
        return
    lst = st.session_state.setdefault(_FT_XLSX_EXTRA_KEY, [])
    if t not in lst:
        lst.append(t)


def _xlsx_remove_extra(ticker: str) -> None:
    """Drop a ticker from the export list."""
    lst = st.session_state.get(_FT_XLSX_EXTRA_KEY, [])
    if ticker in lst:
        lst.remove(ticker)


def _cb_remove_extra_pill() -> None:
    """Pills on_change: remove the added-export ticker whose chip was clicked,
    then reset the selection so each chip acts as a one-shot ✕ remove."""
    picked = st.session_state.get("_sf_xlsx_rm_pills")
    if picked:
        _xlsx_remove_extra(str(picked).split()[0])
    st.session_state["_sf_xlsx_rm_pills"] = None


def _tier_bands_from_tiers(spread_tiers) -> dict:
    return tier_bands(spread_tiers)


def _prior_week_close(conn, ticker: str, week_start: str):
    """Last weekly close strictly before week_start (XSP scaled to /10).

    Reads weekly_spx (SPX/XSP) or weekly_underlying (own-HAR tickers).
    The old exporter read a `spx_close` column off model_features — which
    doesn't exist in that table — so Prev Close was silently blank on
    every export.
    """
    from phase1.ticker_config import uses_own_har, get_config, price_scale_divisor
    try:
        if uses_own_har(ticker):
            from range_finder.data_collector import get_weekly_underlying
            wk = get_weekly_underlying(conn, ticker=ticker)
            col = "close"
        else:
            wk = rf_get_weekly_spx(conn)
            col = "spx_close"
        if wk.empty or col not in wk.columns:
            return None
        prior = wk.loc[wk.index < pd.Timestamp(week_start), col].dropna()
        if prior.empty:
            return None
        val = float(prior.iloc[-1])
        val /= price_scale_divisor(ticker)
        return round(val, 2)
    except Exception:
        return None


@st.cache_data(ttl=600, show_spinner=False)
def _cached_prior_week_close(week_start: str, ticker: str):
    """Cross-session cache so the export build (which runs on every rerun
    because st.download_button materializes its payload eagerly) doesn't
    re-SELECT weekly history each interaction."""
    return _prior_week_close(_get_rf_conn(), ticker, week_start)


def _collect_week_bands_for_ticker(ticker: str, model_choice: str, week_start: str) -> dict:
    """Forecast + tier strikes for one NON-ACTIVE ticker from persisted
    state only: the saved HAR fit, the cron's Monday-open capture
    (weekly_setup), and the DB weekly-EM snapshot. Returns a plain dict
    so st.cache_data can pickle it; failures degrade to an error note on
    that ticker's row instead of sinking the whole export.
    """
    from phase1.ticker_config import feature_source_ticker
    from phase1.gex_history import get_em_snapshot

    out = {"ticker": ticker, "ref": None, "prev_close": None,
           "bands": {}, "notes": [], "error": None}
    try:
        conn = _get_rf_conn()

        # A scaled mini (XSP→SPX) reads its parent's shared features.
        _src = feature_source_ticker(ticker)
        df_feat = _cached_rf_get_features(conn, ticker=_src)
        if df_feat.empty:
            out["error"] = "no feature data — run Weekly Setup on this ticker"
            return out

        # Saved fit for the active spec. A scaled mini shares its parent's fit;
        # the Monday cron may save under either key, so try the ticker first,
        # then fall back to the feature-source parent (identical by construction).
        try:
            payload = _cached_rf_load_model(model_choice, ticker)
        except Exception:
            if _src != ticker:
                try:
                    payload = _cached_rf_load_model(model_choice, _src)
                    out["notes"].append(f"{ticker} via {_src} fit")
                except Exception:
                    out["error"] = f"no saved {model_choice} fit"
                    return out
            else:
                out["error"] = f"no saved {model_choice} fit — run Weekly Setup on {ticker}"
                return out

        if uses_disabled_gex_feature(payload["feature_cols"]):
            out["error"] = (
                f"saved {model_choice} fit uses disabled GEX calibration — "
                f"run Weekly Setup on {ticker}"
            )
            return out

        wk_ts = pd.Timestamp(week_start)
        if wk_ts not in df_feat.index:
            out["error"] = (
                f"forecast blocked: no feature row for {week_start}; "
                "run Weekly Setup after the prior week closes"
            )
            return out
        feature_row = df_feat.loc[wk_ts]
        _fresh, _source_week, _required_week = rf_assess_path_provenance(
            feature_row, wk_ts,
        )
        if not _fresh:
            _source_label = (
                _source_week.strftime("%Y-%m-%d")
                if _source_week is not None else "missing"
            )
            out["error"] = (
                f"forecast blocked: {week_start} path source is "
                f"{_source_label}; requires {_required_week:%Y-%m-%d}"
            )
            return out

        # Reference price: Monday-open capture, else prior weekly close.
        ref = None
        vix = None
        setup = _cached_weekly_setup(conn, week_start, ticker)
        if setup:
            ref = float(setup[0]) if setup[0] else None
            vix = float(setup[1]) if setup[1] else None
        prev_close = _cached_prior_week_close(week_start, ticker)
        if ref is None:
            ref = prev_close
            if ref is not None:
                out["notes"].append("ref = prior close (no Mon-open capture)")
        if ref is None:
            out["error"] = "no reference price (weekly_setup empty)"
            return out
        if vix is None:
            try:
                _v = feature_row.get("vix_close")
                vix = float(_v) if _v is not None and _v == _v else 18.0
            except (TypeError, ValueError):
                vix = 18.0

        try:
            wem = get_em_snapshot(week_start, ticker=ticker, em_type="weekly")
        except Exception:
            wem = None

        # Live Friday chain so the export snaps to the SAME real listed strikes
        # the on-screen tab does — nominal increment rounding otherwise drifts
        # from the tradeable grid (e.g. AMD 556 nominal vs 557.5 on its $2.5
        # far-OTM grid). Degrades to nominal strikes on any failure.
        chain_quotes, chain_exp = _export_chain_quotes(
            ticker, pd.Timestamp(week_start).date()
        )

        _side_q = _cached_side_share_q(ticker)
        forecast, plan, tiers = build_recommendations(
            result=payload["result"], feature_row=feature_row,
            feature_cols=payload["feature_cols"], reference=ref, vix=vix,
            week_start=week_start, ticker=ticker, side_share_q=_side_q,
            chain_quotes=chain_quotes or None, weekly_em=wem,
            conn=conn, model_name=model_choice,
        )

        out["ref"] = round(float(ref), 2)
        out["prev_close"] = prev_close
        out["bands"] = _tier_bands_from_tiers(tiers)

        bits = [model_choice]
        _events = [n for n, f in (("FOMC", plan.has_fomc), ("CPI", plan.has_cpi),
                                  ("NFP", plan.has_nfp), ("OPEX", plan.has_opex)) if f]
        if _events:
            bits.append("/".join(_events))
        if plan.buffer_pct:
            bits.append(f"buf {plan.buffer_pct * 100:.2f}%")
        if plan.recommended_width:
            bits.append(f"wing {plan.recommended_width:g}")
        bits.append(f"chain {chain_exp}" if chain_exp else "nominal strikes")
        out["notes"] = bits + out["notes"]
    except Exception as e:
        out["error"] = f"{type(e).__name__}: {e}"
    return out


@st.cache_data(ttl=600, show_spinner=False)
def _cached_nonactive_week_bands(week_start: str, model_choice: str, tickers: tuple) -> list[dict]:
    return [_collect_week_bands_for_ticker(t, model_choice, week_start) for t in tickers]


def _ft_monday_tab_names(week_start: str) -> list[str]:
    """Every Monday tab name the Scoreboard pre-scans, as ISO text strings.

    Week tabs are named for their planning Monday (yyyy-mm-dd), so the whole
    horizon is enumerable up front: _FT_SB_WEEKS_BACK Mondays behind the
    export week (covers tabs already accumulated in an older master) through
    _FT_SB_WEEKS_FWD ahead. A listed tab that doesn't exist yet is remapped
    by the Scoreboard's hidden column-L sanitizer — IF(ISREF(INDIRECT(…)),
    name, _FT_BLANK_SHEET) — to the shipped empty blank sheet, so every
    INDIRECT target always exists and missing weeks contribute nothing.
    (Deliberately no IFERROR: wrapping the SUMIF-over-INDIRECT array in
    IFERROR collapses the whole SUMPRODUCT to 0 in real Excel — see
    _write_ft_scoreboard.)
    """
    anchor = datetime.strptime(week_start, "%Y-%m-%d").date()
    anchor -= timedelta(days=anchor.weekday())      # snap to Monday
    first = anchor - timedelta(weeks=_FT_SB_WEEKS_BACK)
    return [(first + timedelta(weeks=k)).isoformat()
            for k in range(_FT_SB_WEEKS_BACK + _FT_SB_WEEKS_FWD + 1)]


def _write_ft_scoreboard(sb, week_start: str, tickers: list[str]) -> None:
    """Write the ticker-keyed Scoreboard onto worksheet ``sb``.

    Row-agnostic by design: every formula matches instruments BY NAME
    (column A) against the Instrument column of every week tab, so a ticker
    can sit on a different row — or be absent — on any given week without
    corrupting the tallies. Aggregation is SUMPRODUCT + SUMIF/COUNTIFS over
    INDIRECT references built from the pre-listed Monday tab names (classic
    multi-sheet conditional-sum pattern; no 365-only functions, works in
    LibreOffice too). Excel refuses to array-lift that pattern once IFERROR
    wraps it (verified empirically — the whole product silently collapses
    to 0), so not-yet-pasted weeks CANNOT be skipped via error handling;
    instead hidden column L sanitizes each name with ISREF, mapping missing
    tabs to the shipped, empty _FT_BLANK_SHEET so every INDIRECT target
    always exists. INDIRECT is volatile, so pasting a new week tab re-runs
    the sanitizer and the tab is picked up immediately. Nothing references
    another sheet except through INDIRECT strings, so Move/Copy of this
    sheet (together with _FT_BLANK_SHEET) into an older master carries the
    whole scoring system without creating external links.

    Visible layout: A Instrument · B Weeks Scored · C..J (Wins, Hit %) per
    tier, one row per slot, ALL row at the bottom; K1 self-diagnoses (live
    tab count, duplicate-"(2)"-tab and horizon warnings). Hidden: L sanitized
    week-tab names · M raw pre-listed Monday names · N..Q per-tier
    scored-week denominators · R duplicate-tab detector. A week only enters
    a tier's denominator when that tier's flag is numeric AND its Low band
    cell holds a number — the band criterion keeps pre-rework tabs (whose
    old unguarded flags emit 0 on band-less rows) from scoring false losses.
    The sheet ships protected (no password) with only the Instrument slots
    editable, so the plumbing can't be edited or shifted by accident.
    """
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Protection, Side,
    )

    NAVY, MUTED = "1F3864", "595959"
    HDR_FILL = PatternFill("solid", start_color=NAVY)
    TIER_FILLS = [PatternFill("solid", start_color=c)
                  for c in ("FCE4D6", "FFF2CC", "E2EFDA", "DDEBF7")]
    thin = Side(style="thin", color="BFBFBF")
    BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

    names = _ft_monday_tab_names(week_start)
    names_last = _FT_SB_FIRST_ROW + len(names) - 1

    def _ind(col: str) -> str:
        # e.g. INDIRECT("'"&$L$6:$L$266&"'!$B$6:$B$45") — one range per listed
        # week tab, through the sanitized names in L, so every target exists
        # (missing weeks point at the blank sheet and contribute nothing).
        # Deliberately NO IFERROR: if the blank sheet gets deleted the board
        # shows #REF! loudly instead of silently under-counting.
        return (f'INDIRECT("\'"&$L${_FT_SB_FIRST_ROW}:$L${names_last}&"\'!'
                f'${col}${_FT_FIRST_DATA_ROW}:${col}${_FT_LAST_DATA_ROW}")')

    sb["A1"] = "HAR FORWARD TEST — SCOREBOARD"
    sb["A1"].font = Font(bold=True, size=13, color=NAVY)
    # A2/A3 deliberately NOT merged: unmerged text overflows across the empty
    # cells to its right, while a merged cell clips at its own edge.
    sb["A2"] = (
        "Ticker-keyed: week tabs are matched on their Instrument column, so the "
        "ticker mix AND row order are free to differ week to week. Each week: "
        "download the new export, right-click its week tab → Move or Copy → into "
        "this workbook, any position. One tab per Monday — re-importing a week? "
        "Delete its old tab FIRST (Excel silently renames the paste to (2), "
        "which the board ignores)."
    )
    sb["A2"].font = Font(italic=True, size=9, color=MUTED)
    sb["A3"] = (
        f"Track a new instrument any time: type its ticker into a blank "
        f"Instrument cell — every week tab in the book re-scores instantly. "
        f"Don't rename week tabs or insert rows/columns on them; keep the "
        f"'{_FT_BLANK_SHEET}' sheet. Hit % counts only weeks where that tier "
        f"had a band and a filled close. Weeks after {names[-1]} need a fresh "
        f"export's Scoreboard (a warning appears here in time). Upgrading an "
        f"old master: also delete its WeeksStart/WeeksEnd tabs and re-type its "
        f"extra tickers here."
    )
    sb["A3"].font = Font(italic=True, size=9, color=MUTED)

    # Self-diagnostics: live week-tab count, plus loud warnings for the two
    # silent failure modes — a pasted duplicate that Excel renamed to "(2)"
    # (invisible to the name-keyed scan) and the pre-listed name horizon
    # running out.
    _horizon_end = date_cls.fromisoformat(names[-1])
    _warn_from = _horizon_end - timedelta(weeks=12)
    sb["K1"] = (
        f'="Week tabs found: "'
        f'&SUMPRODUCT(--($L${_FT_SB_FIRST_ROW}:$L${names_last}<>"{_FT_BLANK_SHEET}"))'
        f'&IF(SUM($R${_FT_SB_FIRST_ROW}:$R${names_last})>0,'
        f'"   ⚠ a pasted week tab was renamed like {week_start} (2) — the board '
        f'ignores it; delete the old tab, then rename the (2) tab to the plain '
        f'date","")'
        f'&IF(TODAY()>DATE({_warn_from.year},{_warn_from.month},{_warn_from.day}),'
        f'"   ⚠ last pre-listed week is {names[-1]} — swap in a fresh export '
        f'Scoreboard + {_FT_BLANK_SHEET}","")'
    )
    sb["K1"].font = Font(italic=True, size=9, color=MUTED)

    _sb_groups = [("C", "D", "LOWER PI"), ("E", "F", "POINT EST"),
                  ("G", "H", "80% PI UPPER"), ("I", "J", "EFFECTIVE")]
    for (lo, hi, label), fill in zip(_sb_groups, TIER_FILLS):
        sb.merge_cells(f"{lo}4:{hi}4")
        c = sb[f"{lo}4"]
        c.value = label
        c.fill = fill
        c.font = Font(bold=True, size=9)
        c.alignment = CENTER
    for col, label in [("A", "Instrument"), ("B", "Weeks Scored")] + [
        (c, lbl) for grp in _sb_groups for c, lbl in ((grp[0], "Wins"), (grp[1], "Hit %"))
    ]:
        cell = sb[f"{col}5"]
        cell.value = label
        cell.fill = HDR_FILL
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.alignment = CENTER
        cell.border = BOX
    for col, label in [("L", "Week tabs (live)"), ("M", "Week tabs (all)"),
                       ("N", "n LPI"), ("O", "n Point"),
                       ("P", "n 80PI"), ("Q", "n Eff"), ("R", "dup?")]:
        cell = sb[f"{col}5"]
        cell.value = label
        cell.font = Font(bold=True, size=8, color=MUTED)

    # M: pre-listed Monday tab names, stored as TEXT (INDIRECT concatenates
    # them into sheet references — Excel must never coerce one to a date
    # serial). L: the sanitized live name every formula actually uses — the
    # week tab itself when it exists, the blank sheet when it doesn't.
    # R: flags a duplicate tab Excel renamed to "<Monday> (2)" (feeds K1).
    for k, nm in enumerate(names):
        r = _FT_SB_FIRST_ROW + k
        c = sb[f"M{r}"]
        c.number_format = "@"
        c.value = nm
        sb[f"L{r}"] = (f'=IF(ISREF(INDIRECT("\'"&$M{r}&"\'!$A$1")),'
                       f'$M{r},"{_FT_BLANK_SHEET}")')
        sb[f"R{r}"] = f'=IF(ISREF(INDIRECT("\'"&$M{r}&" (2)\'!$A$1")),1,0)'

    # Instrument slots: export tickers first, the rest blank-but-armed — every
    # slot carries the full formula set, guarded on column A being non-blank
    # (IF short-circuits, so empty slots cost nothing to recalc). Slot cells
    # are the only unlocked cells on the sheet (typing a ticker must survive
    # the sheet protection below).
    # (win col, hidden denominator col, week-tab flag col, band-Low col):
    _win_cols = [("C", "N", "O", "G"), ("E", "O", "P", "I"),
                 ("G", "P", "Q", "K"), ("I", "Q", "R", "M")]
    for idx in range(_FT_SB_SLOTS):
        r = _FT_SB_FIRST_ROW + idx
        tick = f"$A{r}"
        if idx < len(tickers):
            sb[f"A{r}"] = _xlsx_safe_cell(tickers[idx])
            sb[f"A{r}"].font = Font(bold=True)
        sb[f"A{r}"].protection = Protection(locked=False)
        sb[f"B{r}"] = (f'=IF({tick}="","",SUMPRODUCT('
                       f'SUMIF({_ind("B")},{tick},{_ind("S")})))')
        for win_col, den_col, flag_col, lo_col in _win_cols:
            sb[f"{win_col}{r}"] = (
                f'=IF({tick}="","",SUMPRODUCT('
                f'COUNTIFS({_ind("B")},{tick},{_ind(flag_col)},1)))'
            )
            # Denominator: flags are numeric (0/1) only when the tier scored;
            # ">=0" skips the ""-blanks bandless/unfilled rows produce. The
            # extra band-Low ">0" criterion drops PRE-REWORK rows whose old
            # unguarded flags emit 0 with no band at all (false losses).
            sb[f"{den_col}{r}"] = (
                f'=IF({tick}="","",SUMPRODUCT('
                f'COUNTIFS({_ind("B")},{tick},{_ind(flag_col)},">=0",'
                f'{_ind(lo_col)},">0")))'
            )
        for hit_col, (win_col, den_col, *_cols) in zip("DFHJ", _win_cols):
            sb[f"{hit_col}{r}"] = (
                f'=IF({tick}="","",IF({den_col}{r}=0,"—",{win_col}{r}/{den_col}{r}))'
            )
            sb[f"{hit_col}{r}"].number_format = "0%"
        for col in "ABCDEFGHIJ":
            sb[f"{col}{r}"].border = BOX
            if col != "A":
                sb[f"{col}{r}"].alignment = Alignment(horizontal="center")

    # ALL row — pooled across every instrument slot.
    ar = _FT_SB_ALL_ROW
    lo_r, hi_r = _FT_SB_FIRST_ROW, _FT_SB_LAST_ROW
    sb[f"A{ar}"] = "ALL"
    sb[f"A{ar}"].font = Font(bold=True)
    sb[f"B{ar}"] = f"=SUM(B{lo_r}:B{hi_r})"
    for win_col, den_col, *_cols in _win_cols:
        sb[f"{win_col}{ar}"] = f"=SUM({win_col}{lo_r}:{win_col}{hi_r})"
        sb[f"{den_col}{ar}"] = f"=SUM({den_col}{lo_r}:{den_col}{hi_r})"
    for hit_col, (win_col, den_col, *_cols) in zip("DFHJ", _win_cols):
        sb[f"{hit_col}{ar}"] = f'=IF({den_col}{ar}=0,"—",{win_col}{ar}/{den_col}{ar})'
        sb[f"{hit_col}{ar}"].number_format = "0%"
    for col in "ABCDEFGHIJ":
        sb[f"{col}{ar}"].border = BOX
        if col != "A":
            sb[f"{col}{ar}"].alignment = Alignment(horizontal="center")

    sb.column_dimensions["A"].width = 12
    sb.column_dimensions["B"].width = 13
    for col in "CDEFGHIJ":
        sb.column_dimensions[col].width = 9
    for col in "LMNOPQR":
        sb.column_dimensions[col].hidden = True
    sb.freeze_panes = "A6"

    # Guardrail: protect the sheet (no password) so the hidden plumbing and
    # formulas can't be edited or shifted by accident — only the Instrument
    # slots (unlocked above) accept typing. Formatting stays allowed;
    # right-click the tab → Unprotect Sheet to override deliberately.
    sb.protection.sheet = True
    sb.protection.formatCells = False
    sb.protection.formatColumns = False
    sb.protection.formatRows = False


def _write_ft_week_sheet(wb, week_start: str, rows: list[dict]):
    """Write one week tab (named for its Monday) into ``wb``; returns it.

    Self-contained on purpose: every formula references THIS sheet only, so
    Move/Copy into the master never externalizes a reference. The Scoreboard
    matches its rows by the Instrument column, so ``rows`` may carry any
    ticker mix, in any order (capped at _FT_MAX_TICKERS — the Scoreboard
    scans B6:B45). Every row of that window is written: exported tickers on
    top, the rest blank-but-armed (formulas + styling in place) so a row the
    user types in by hand — the only way to log the current week for a
    freshly tracked ticker — scores exactly like an exported one. Ships
    protected (no password): data cells editable, formula cells and the
    sheet structure locked (the Scoreboard reads fixed positions).
    """
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Protection, Side,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    NAVY, MUTED = "1F3864", "595959"
    HDR_FILL = PatternFill("solid", start_color=NAVY)
    TIER_FILLS = [PatternFill("solid", start_color=c)
                  for c in ("FCE4D6", "FFF2CC", "E2EFDA", "DDEBF7")]
    GREY_FILL = PatternFill("solid", start_color="D9D9D9")
    GREEN_FILL = PatternFill("solid", start_color="C6EFCE")
    RED_FILL = PatternFill("solid", start_color="FFC7CE")
    thin = Side(style="thin", color="BFBFBF")
    BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rows = list(rows)[:_FT_MAX_TICKERS]
    FIRST = _FT_FIRST_DATA_ROW
    ws = wb.create_sheet(week_start)

    ws["A1"] = f"HAR MODELS — WEEKLY RANGE FORWARD TEST — week of {week_start}"
    ws["A1"].font = Font(bold=True, size=13, color=NAVY)
    ws["A2"] = (
        "One row per instrument. Each model tier gives its own Low/High band. "
        "Win = weekly CLOSE inside that tier's band (intraday wicks ignored — set-and-forget). "
        "Fill 'Weekly Close' (column E) after Friday's close; CLOSE INSIDE, Scored? and the "
        "Scoreboard update automatically. The Scoreboard matches rows by Instrument name, "
        "so row order doesn't matter — just don't rename this tab or insert rows/columns "
        "(sheet is protected, no password, to keep the scored layout fixed; all data "
        "cells stay editable — add extra instrument rows below the exported ones freely)."
    )
    ws["A2"].font = Font(italic=True, size=9, color=MUTED)
    ws.merge_cells("A2:T2")

    _tier_groups = [
        ("G", "H", "LOWER PI (tightest — richest credit)"),
        ("I", "J", "POINT ESTIMATE (HAR base)"),
        ("K", "L", "80% PI UPPER (80% interval)"),
        ("M", "N", "EFFECTIVE (PI + buffer — your condor)"),
    ]
    for (lo, hi, label), fill in zip(_tier_groups, TIER_FILLS):
        ws.merge_cells(f"{lo}4:{hi}4")
        c = ws[f"{lo}4"]
        c.value = label
        c.fill = fill
        c.font = Font(bold=True, size=8)
        c.alignment = CENTER
    ws.merge_cells("O4:R4")
    ws["O4"] = "CLOSE INSIDE? (per model)"
    ws["O4"].fill = PatternFill("solid", start_color="B4C6E7")
    ws["O4"].font = Font(bold=True, size=8)
    ws["O4"].alignment = CENTER

    _headers = [
        ("A", "Week (Mon)"), ("B", "Instrument"), ("C", "Class"),
        ("D", "Ref/Open ($)"), ("E", "Weekly Close ($)"), ("F", "Prev Close ($)"),
        ("G", "Low"), ("H", "High"), ("I", "Low"), ("J", "High"),
        ("K", "Low"), ("L", "High"), ("M", "Low"), ("N", "High"),
        ("O", "Lower PI"), ("P", "Point"), ("Q", "80% PI"), ("R", "Eff"),
        ("S", "Scored?"), ("T", "Notes"),
    ]
    for col, label in _headers:
        cell = ws[f"{col}5"]
        cell.value = label
        cell.fill = HDR_FILL
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.alignment = CENTER
        cell.border = BOX

    _slot_cols = [("lower_pi", "G", "H"), ("point", "I", "J"),
                  ("pi_upper", "K", "L"), ("effective", "M", "N")]
    _week_date = None
    try:
        _week_date = datetime.strptime(week_start, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        pass

    # Every scannable row (6..45) is written: exported tickers on top, the
    # rest blank-but-armed so a row the user types in by hand scores exactly
    # like an exported one (the Scoreboard reads the whole window either way).
    for r in range(FIRST, _FT_LAST_DATA_ROW + 1):
        row = rows[r - FIRST] if r - FIRST < len(rows) else None
        if row is not None:
            t = row.get("ticker", "?")
            ws[f"A{r}"] = _week_date or week_start
            ws[f"A{r}"].number_format = "yyyy-mm-dd"
            ws[f"B{r}"] = _xlsx_safe_cell(t)
            ws[f"B{r}"].font = Font(bold=True)
            ws[f"C{r}"] = _ft_class(t)
            if row.get("ref") is not None:
                ws[f"D{r}"] = row["ref"]
            if row.get("prev_close") is not None:
                ws[f"F{r}"] = row["prev_close"]
            bands = row.get("bands") or {}
            for slot, lo_col, hi_col in _slot_cols:
                band = bands.get(slot)
                if band is not None:
                    ws[f"{lo_col}{r}"] = band[0]
                    ws[f"{hi_col}{r}"] = band[1]
            note = " · ".join(row.get("notes") or [])
            if row.get("error"):
                note = (f"⚠ {row['error']}" + (f" · {note}" if note else ""))
            ws[f"T{r}"] = note
            ws[f"T{r}"].font = Font(size=8, color=MUTED)
            ws[f"T{r}"].alignment = Alignment(wrap_text=True, vertical="center")

        # E — Weekly Close: left blank for the user, shaded as fill-me.
        ws[f"E{r}"].fill = GREY_FILL
        for (slot, lo_col, hi_col), fill in zip(_slot_cols, TIER_FILLS):
            ws[f"{lo_col}{r}"].fill = fill
            ws[f"{hi_col}{r}"].fill = fill

        # CLOSE INSIDE? flags — blank until the row has an instrument AND the
        # close AND that tier's band are all NUMBERS (COUNT counts numbers
        # only): a pasted close with a stray space or an 'n/a' band stays
        # blank instead of scoring a false loss (any text compares above any
        # number in Excel, so an unguarded compare emits 0). 1/0 once
        # scoreable — the Scoreboard's denominators count exactly these
        # numeric cells.
        for (slot, lo_col, hi_col), flag_col in zip(_slot_cols, "OPQR"):
            ws[f"{flag_col}{r}"] = (
                f'=IF(OR($B{r}="",COUNT($E{r},{lo_col}{r},{hi_col}{r})<3),"",'
                f'IF(AND($E{r}>={lo_col}{r},$E{r}<={hi_col}{r}),1,0))'
            )
            ws[f"{flag_col}{r}"].alignment = Alignment(horizontal="center")
        # Scored? = instrument present, close is a number, ≥1 numeric band.
        ws[f"S{r}"] = (f'=IF(OR($B{r}="",COUNT($E{r})=0,'
                       f'COUNT($G{r}:$N{r})=0),"",1)')
        ws[f"S{r}"].alignment = Alignment(horizontal="center")

        for col_idx in range(1, 21):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = BOX
            if 4 <= col_idx <= 14 and not isinstance(cell.value, str):
                cell.number_format = "#,##0.00"

    # Green/red the CLOSE INSIDE cells once scored — the whole armed window,
    # so hand-added rows colour like exported ones.
    ws.conditional_formatting.add(
        f"O{FIRST}:R{_FT_LAST_DATA_ROW}",
        CellIsRule(operator="equal", formula=["1"], fill=GREEN_FILL),
    )
    ws.conditional_formatting.add(
        f"O{FIRST}:R{_FT_LAST_DATA_ROW}",
        CellIsRule(operator="equal", formula=["0"], fill=RED_FILL),
    )

    _widths = {"A": 11, "B": 11, "C": 7, "D": 10, "E": 11, "F": 10,
               "S": 8, "T": 46}
    for col_idx in range(7, 15):
        _widths[get_column_letter(col_idx)] = 9
    for col_idx in range(15, 19):
        _widths[get_column_letter(col_idx)] = 8
    for col, w in _widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"

    # Weekly Close must be a number — typed text silently un-scores the row
    # (the COUNT guards keep flags blank rather than emit a false loss), so
    # also reject it at entry. Paste bypasses validation; the formula guards
    # stay authoritative.
    dv = DataValidation(
        type="decimal", operator="greaterThan", formula1="0", allowBlank=True,
        showErrorMessage=True, errorTitle="Weekly Close must be a number",
        error=("Type the Friday close as a plain number — text (even a stray "
               "space) keeps the row from scoring."),
    )
    ws.add_data_validation(dv)
    dv.add(f"E{FIRST}:E{_FT_LAST_DATA_ROW}")

    # Guardrail: protect the sheet (no password). Data-entry cells stay
    # unlocked (B..N instrument/class/ref/close/bands + T notes — enough to
    # hand-add a full instrument row); the flag/Scored? formulas and the
    # sheet STRUCTURE stay locked, because the Scoreboard reads fixed
    # positions (B6:B45, flags O..R, Scored? S) and a single inserted
    # row/column would silently garble every tally. Formatting stays
    # allowed; right-click the tab → Unprotect Sheet to override.
    for r in range(FIRST, _FT_LAST_DATA_ROW + 1):
        for col in "BCDEFGHIJKLMNT":
            ws[f"{col}{r}"].protection = Protection(locked=False)
    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    return ws


def _write_ft_blank_sheet(wb) -> None:
    """Create the empty utility sheet unpasted weeks resolve to.

    The Scoreboard's sanitizer (hidden column L) points every pre-listed
    Monday whose tab isn't in the workbook yet at this sheet, so its scanned
    region (rows 6-45) must stay empty — a note in row 1 is safely outside.
    """
    from openpyxl.styles import Font

    bk = wb.create_sheet(_FT_BLANK_SHEET)
    bk["A1"] = (
        f"Keep this sheet (and keep it empty) — the Scoreboard resolves "
        f"not-yet-added weeks here. If you copy the Scoreboard into another "
        f"workbook, copy '{_FT_BLANK_SHEET}' along with it. The sheet is "
        f"protected on purpose: anything typed into it would be counted once "
        f"per missing week."
    )
    bk["A1"].font = Font(italic=True, size=9, color="595959")
    bk.sheet_view.showGridLines = False
    # Fully locked, no password: a stray row typed here would be multiplied
    # by every not-yet-pasted Monday (~150-200x) with no error anywhere.
    bk.protection.sheet = True


def _build_forward_test_workbook(*, week_start: str, model_choice: str, rows: list[dict]) -> bytes:
    """Assemble the multi-ticker forward-test workbook.

    Sheets, in order: Scoreboard · <week_start> · FTBlank. `rows` is the
    ordered list of per-ticker dicts to write (defaults first, then
    user-added extras); each carries its own ``ticker`` (see
    _collect_week_bands_for_ticker for the shape). The Scoreboard matches
    week tabs by ticker name, so the mix and order may differ freely across
    weeks (capped at _FT_MAX_TICKERS per week — the fixed range its formulas
    scan). ``model_choice`` only flavors the notes the caller baked into
    ``rows``.
    """
    from io import BytesIO

    import openpyxl

    rows = list(rows)[:_FT_MAX_TICKERS]
    wb = openpyxl.Workbook()
    sb = wb.active
    sb.title = "Scoreboard"
    # Slot labels feed SUMIF/COUNTIFS criteria directly — drop blanks rather
    # than emit a placeholder ("?" is an Excel wildcard matching any 1-char
    # instrument).
    _slot_tickers = [t for t in (str(r.get("ticker") or "").strip()
                                 for r in rows) if t]
    _write_ft_scoreboard(sb, week_start, _slot_tickers)
    _write_ft_week_sheet(wb, week_start, rows)
    _write_ft_blank_sheet(wb)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _auto_warm_up_spread_model(conn, ticker: str, ticker_cfg: dict) -> bool:
    """Cold-start the weekly Spread Finder for a freshly-looked-up ticker.

    Mirrors the Refresh → Rebuild → Forecast button chain but runs
    automatically the first time an own-HAR ticker has no feature matrix and
    no fitted model. Pulls ~6yr of weekly OHLC from yfinance, builds the
    feature matrix, then fits + saves every model spec to Postgres so later
    loads are instant.

    Returns True when a usable feature matrix now exists; False (caller shows a
    graceful "unavailable" note) when yfinance has no history for the symbol.
    """
    from phase1.ticker_config import uses_own_har as _uses_own_har
    from phase1.ticker_config import feature_source_ticker as _feature_source

    feat_ticker = _feature_source(ticker)

    # 1) Per-ticker weekly OHLC (own-HAR only; a scaled mini rides its parent's
    #    rows, which the parent's own warm-up/cron populates).
    if _uses_own_har(ticker):
        try:
            from range_finder.data_collector import (
                fetch_underlying_weekly, save_underlying_weekly,
            )
            df_t = fetch_underlying_weekly(
                ticker=ticker,
                yf_symbol=ticker_cfg["yf_symbol"],
                vol_proxy_yf=ticker_cfg.get("vol_proxy_yf", "^VIX"),
                years=6,
            )
            if df_t is None or len(df_t) == 0:
                return False
            save_underlying_weekly(conn, ticker, df_t)
        except Exception as e:
            st.caption(f"⚠ Could not load price history for {ticker}: {e}")
            return False

    # 2) Build the feature matrix.
    try:
        rf_build_features(conn, ticker=feat_ticker)
        _cached_rf_get_features.clear()
    except Exception as e:
        st.caption(f"⚠ Feature build failed for {ticker}: {e}")
        return False

    # 3) Fit + save every spec (same logic as the Weekly Setup path).
    try:
        df_feat = _cached_rf_get_features(conn, ticker=feat_ticker)
    except Exception:
        df_feat = pd.DataFrame()
    if df_feat is None or df_feat.empty:
        return False

    fitted_any = False
    for _spec in RF_MODEL_SPECS.keys():
        feat_cols = live_spread_feature_columns(RF_MODEL_SPECS[_spec])
        if _spec == "M4_full" and GEX_LIVE_SPREAD_INFLUENCE_ENABLED:
            gex_col = GEX_NORMALIZED_FEATURE
            if rf_feature_has_enough_data(df_feat, gex_col) and gex_col not in feat_cols:
                feat_cols.append(gex_col)
        avail_cols = production_feature_columns(df_feat, _spec)
        if len(avail_cols) < 2:
            continue
        try:
            _validation, _result, _metrics = rf_fit_validation_and_production(
                df_feat, feature_cols=avail_cols, model_name=_spec
            )
            rf_save_model(_result, avail_cols, _spec, _metrics,
                          conn=conn, ticker=ticker)
            fitted_any = True
        except Exception:
            continue
    _cached_rf_load_model.clear()
    return fitted_any


@st.fragment
def _render_spread_finder_tab(spot: float, levels: dict, regime: dict, data, ticker: str = "SPX", weekly_em: dict = None):
    """Render HAR spread placement with GEX kept as research context.

    Wrapped in @st.fragment so widget interactions inside the tab (horizon
    slider, model-spec dropdown, credit width, etc.) only rerun this tab
    instead of triggering a full-page rerun that rebuilds the GEX chart,
    the sidebar, and re-fetches weekly/monthly EM."""
    import yfinance as yf
    from phase1.market_clock import now_ny

    # resolve_config returns the chain-derived config registered by the main
    # app for an arbitrary ticker (or the curated entry for the known five) —
    # never SPX's 5-point grid as a silent fallback for an unknown symbol.
    from phase1.ticker_config import resolve_config as _resolve_config
    ticker_cfg = _resolve_config(ticker)

    # ── Earnings-week warning (single-stock tickers only) ──
    # AMZN/AMD set has_single_name_earnings=True in ticker_config; QQQ/SPX/XSP
    # don't. When the upcoming week is flagged in earnings_flags, surface a
    # warning banner at the top of the tab but still let the user generate
    # a plan — single-stock IV is structurally elevated pre-earnings, which
    # makes weekly OTM credit spreads look juicy but their tail risk is
    # much higher than the HAR predicts. ("Warn but allow" mode.)
    from phase1.ticker_config import has_single_name_earnings as _has_sne
    if _has_sne(ticker):
        try:
            from datetime import timedelta as _td_e
            from phase1.market_clock import now_ny as _now_ny_e
            # Check the SAME week the spread finder is planning for —
            # this week's Monday on Mon-Thu, next Monday on Fri-Sun
            # (mirrors _spread_finder_target_friday / week_start below).
            # The old check always looked at NEXT Monday, so a Tuesday
            # user planning this week's spreads never saw the warning
            # for earnings landing on Wednesday.
            #
            # Use the EXCHANGE clock (now_ny), not server-local: a UTC-hosted
            # Streamlit Cloud instance rolls to the next calendar day at
            # 20:00 ET, so a late-evening ET user would otherwise be shown
            # the wrong planning week (and miss/false-fire the warning) for
            # 4 hours every day. Everything else in the tab keys off ET.
            _today = _now_ny_e().date()
            _wd_e = _today.weekday()
            if _wd_e <= 3:
                _plan_monday = _today - _td_e(days=_wd_e)
            else:
                _plan_monday = _today + _td_e(days=7 - _wd_e)
            _plan_week = _plan_monday.strftime("%Y-%m-%d")
            _conn_e = _get_rf_conn()
            _cur_e = _conn_e.cursor()
            _cur_e.execute(
                "SELECT has_earnings, earnings_date FROM earnings_flags "
                "WHERE ticker = ? AND week_start = ?",
                (ticker, _plan_week),
            )
            _row_e = _cur_e.fetchone()
            if _row_e and _row_e[0]:
                _ed = _row_e[1] or _plan_week
                st.warning(
                    f"⚠ **{ticker} reports earnings the week of {_plan_week}** "
                    f"(scheduled {_ed}). Single-stock IV is structurally elevated "
                    "pre-earnings — weekly OTM credit spreads will look richer "
                    "than usual, but realised tail risk is much higher than the "
                    "HAR model predicts. Verify chain liquidity, and consider "
                    "sizing down or skipping this week."
                )
        except Exception:
            # Don't let an earnings-flag query failure block the tab render.
            pass

    # ── Fetch latest vol-proxy close (5-min TTL, refreshed across reruns) ──
    # Per-ticker vol proxy: VIX for SPX/XSP/AMZN/AMD, VXN for QQQ.
    #   live_vix_raw : the genuine live close, or None when unavailable. This
    #                  is the ONLY value the regime-shift breaker may read — a
    #                  fabricated fallback would silently disable it.
    #   live_vix     : display/default fallback for the VIX input, where a
    #                  concrete number is required so the widget renders.
    _vol_proxy = ticker_cfg.get("vol_proxy_yf", "^VIX")
    live_vix_raw = _live_vol_close(_vol_proxy)
    live_vix = live_vix_raw if live_vix_raw is not None else 18.0

    # ── Monday open freeze logic ──
    # On the weekly freeze day (Monday or Tue after holiday) at market open,
    # capture the true daily-candle Open as the weekly reference. Rest of
    # the week uses that frozen value. Before Monday open (weekends), use
    # the live spot (Friday close).
    run_now = now_ny()
    is_freeze_day = _is_weekly_freeze_day(run_now)
    is_market_open = data.market_open

    mon_open_key = f"sf_monday_open_{ticker}"
    mon_vix_key = f"sf_monday_vix_{ticker}"
    mon_open_week_key = f"sf_monday_open_week_{ticker}"

    # Determine which week we're in (use ISO week number)
    current_week = run_now.isocalendar()[1]

    def _daily_open_today(symbol: str):
        """Today's daily-candle Open, or None if Yahoo hasn't published
        it yet (happens briefly after 9:30 while the first tick settles)."""
        try:
            hist = yf.Ticker(symbol).history(period="5d")
        except Exception:
            return None
        if hist is None or hist.empty or "Open" not in hist.columns:
            return None
        today = run_now.date()
        for ts, row in hist.iterrows():
            if hasattr(ts, "date") and ts.date() == today:
                op = row.get("Open")
                if op is not None and not (isinstance(op, float) and op != op):
                    return float(op)
        return None

    # Freeze Monday's open on the freeze day when market is open
    if is_freeze_day and is_market_open:
        stored_week = st.session_state.get(mon_open_week_key)
        if stored_week != current_week:
            # First market-hours refresh on the freeze day — lock the
            # TRUE daily-candle Open, not whatever tick `spot` happens to
            # land on while this refresh is running. Fall back to spot
            # only if yfinance hasn't returned today's bar yet.
            #
            # Per-ticker yfinance symbols come from ticker_config:
            # SPX/XSP read ^SPX (XSP scales /10); QQQ reads QQQ; AMZN/AMD
            # read their own symbols. Vol proxy is VIX for everyone except
            # QQQ, which uses VXN.
            from phase1.ticker_config import price_scale_divisor as _scale_div_fn
            _yf_underlying = ticker_cfg.get("yf_symbol", "^GSPC")
            _yf_vol_proxy  = ticker_cfg.get("vol_proxy_yf", "^VIX")
            _scale_div     = _scale_div_fn(ticker)

            underlying_daily_open = _daily_open_today(_yf_underlying)
            if underlying_daily_open is not None:
                # A scaled mini (XSP→SPX) reads its parent's symbol and
                # divides by its scale_divisor; everyone else divides by 1.0.
                frozen_spot = round(underlying_daily_open / _scale_div, 2)
            else:
                frozen_spot = round(spot, 2)

            vol_proxy_daily_open = _daily_open_today(_yf_vol_proxy)
            frozen_vix_val = round(vol_proxy_daily_open, 2) if vol_proxy_daily_open is not None else live_vix

            st.session_state[mon_open_key] = frozen_spot
            st.session_state[mon_vix_key] = frozen_vix_val
            st.session_state[mon_open_week_key] = current_week

    # Determine the reference price/VIX and their source label.
    #
    # The Monday-open anchor only applies while we're planning THIS week's
    # spreads (Mon-Thu). From Friday on, the planner targets NEXT week
    # (see week_start below), and anchoring next week's strikes to the
    # *current* week's Monday open carries up to a full week of drift —
    # the best available proxy for next Monday's open is simply the
    # latest price (Friday's close over the weekend).
    _planning_this_week = run_now.weekday() <= 3
    frozen_open = st.session_state.get(mon_open_key)
    frozen_vix = st.session_state.get(mon_vix_key)
    frozen_week = st.session_state.get(mon_open_week_key)
    self_heal_msg = None

    if _planning_this_week and frozen_week == current_week and frozen_open:
        default_ref = frozen_open
        default_vix = frozen_vix or live_vix
        ref_source = "Mon open (frozen)"
    else:
        # Mid-week (Mon-Thu) the strikes must stay anchored to THIS week's
        # Monday open even when the in-session freeze was lost (a fresh browser
        # session drops session state). Two-step recovery:
        #   1) restore the Monday open + VIX from weekly_setup, and
        #   2) if that row is missing (the Monday cron never ran, or this is a
        #      ticker the cron doesn't fit), capture it RETROACTIVELY from
        #      yfinance and persist it — so we lock to Monday instead of
        #      silently drifting on live spot for the rest of the week.
        # _cached_weekly_setup wraps the SELECT in a 15 min cross-session cache
        # so widget reruns / auto-refresh ticks don't keep firing at Neon.
        restored_open = None
        restored_vix = None
        self_healed = False
        if _planning_this_week:
            from datetime import timedelta as _td
            this_monday = (run_now - _td(days=run_now.weekday())).date()
            week_start_str = this_monday.strftime("%Y-%m-%d")
            rf_conn = _get_rf_conn()
            # 1) Restore a previously-captured Monday anchor.
            try:
                cached_setup = _cached_weekly_setup(rf_conn, week_start_str, ticker)
                if cached_setup is not None:
                    restored_open, restored_vix = cached_setup
            except Exception:
                pass
            # 2) Self-heal a missing capture. No spot_fallback on purpose — we
            #    only persist a TRUE Monday daily-candle Open, never a mid-week
            #    live price masquerading as the weekly anchor. If yfinance has
            #    no Monday bar, the capture raises and we fall through to the
            #    live-spot display for the day (nothing bad gets persisted).
            if not restored_open:
                try:
                    healed_open, healed_vix, _src = rf_capture_monday_anchor(
                        rf_conn, ticker, week_start_str, this_monday,
                        spot_fallback=None, live_vix_fallback=live_vix,
                        cfg=ticker_cfg,
                    )
                    if healed_open:
                        restored_open, restored_vix = healed_open, healed_vix
                        self_healed = True
                        _cached_weekly_setup.clear()
                        self_heal_msg = (
                            f"ℹ️ No Monday-open capture existed for the week of "
                            f"{week_start_str} (the Monday setup job didn't run for "
                            f"{ticker}), so the anchor was captured retroactively "
                            f"from Monday's daily bar — strikes are now locked to it "
                            f"for the week. Run **Weekly Setup** if this recurs."
                        )
                except Exception:
                    pass
            # Mirror whatever we landed on into the session freeze so labels read
            # "Mon open" and later reruns skip the DB / yfinance round-trip.
            if restored_open:
                st.session_state[mon_open_key] = restored_open
                if restored_vix:
                    st.session_state[mon_vix_key] = restored_vix
                st.session_state[mon_open_week_key] = current_week

        if restored_open:
            default_ref = restored_open
            default_vix = restored_vix or live_vix
            ref_source = "Mon open (self-healed)" if self_healed else "Mon open (from DB)"
        else:
            default_ref = round(spot, 2)
            default_vix = live_vix
            if run_now.weekday() >= 5:
                ref_source = "Fri close (next-week plan)"
            elif run_now.weekday() == 4:
                ref_source = "live spot (next-week plan)"
            else:
                ref_source = "live spot"

    # ── Reference-price sanity guard ──
    # The ref is sticky (keyed session state, frozen Monday captures, a DB
    # restore, manual edits), so a single bad write anchors every strike
    # in the tab to a bogus level for the rest of the session — e.g. an
    # SPX ref of 100 against a ~7,400 spot walks the whole ladder to the
    # chain edge and produces nonsense spreads. Live spot is always in
    # hand here, so validate against it: hard-correct values that can only
    # be data corruption (no weekly market move is that big), and surface
    # a warning for large-but-conceivable gaps so the user double-checks.
    _REF_RESET_DEV = 0.25   # >25% off spot: corruption, not a market move
    _REF_WARN_DEV  = 0.12   # >12%: extreme — keep it, but make the user look

    def _ref_deviation(v) -> "float | None":
        """|v/spot - 1|, or None when v isn't a usable price."""
        if not spot or spot <= 0:
            return None
        try:
            v = float(v)
        except (TypeError, ValueError):
            return None
        if v <= 0:
            return None
        return abs(v / spot - 1.0)

    ref_guard_msg = None
    if spot and spot > 0:
        _dev_default = _ref_deviation(default_ref)
        if _dev_default is None or _dev_default > _REF_RESET_DEV:
            default_ref = round(spot, 2)
            ref_source = "live spot (auto-corrected)"

    # ── Auto-update reference price and VIX when ticker changes ──
    # Also evict the ticker we're *leaving*'s cached HAR fit so the
    # incoming ticker lands on a clean load-from-Postgres path. Without
    # this, switching SPX→XSP→(GEX tab)→SPX could leave the outgoing
    # ticker's `_mdl_result_{ticker}` / `_mdl_name_{ticker}` session
    # entries in place; the name check at the bottom of this function
    # then sees cached_name == current dropdown choice and skips the
    # reload, freezing the tab on the last-displayed model.
    ref_key = f"sf_ref_price_{ticker}"
    vix_key = f"sf_vix_level_{ticker}"
    prev_ticker = st.session_state.get("_sf_prev_ticker")
    if prev_ticker != ticker:
        st.session_state[ref_key] = default_ref
        st.session_state[vix_key] = default_vix
        if prev_ticker:
            for _suffix in ("sf_model_result_", "sf_model_features_",
                            "sf_model_metrics_", "sf_model_name_"):
                st.session_state.pop(f"{_suffix}{prev_ticker}", None)
        st.session_state["_sf_prev_ticker"] = ticker

    # Also update the defaults on first render if not yet set
    if ref_key not in st.session_state:
        st.session_state[ref_key] = default_ref
    if vix_key not in st.session_state:
        st.session_state[vix_key] = default_vix

    # Validate whatever the session is actually carrying (stale seed from a
    # broken feed, fat-fingered edit, corrupt weekly_setup restore) — the
    # default-ref heal above can't see a value written on an earlier rerun.
    if spot and spot > 0:
        _cur_ref = st.session_state.get(ref_key)
        _dev_cur = _ref_deviation(_cur_ref)
        if _dev_cur is None or _dev_cur > _REF_RESET_DEV:
            ref_guard_msg = (
                f"⚠️ The stored {ticker} reference (`{_cur_ref}`) was "
                f"{'unusable' if _dev_cur is None else f'{_dev_cur:.0%} away from live spot'} "
                f"(spot ≈ {spot:,.2f}) — strikes built from it would be nonsense, "
                f"so it was reset to **{default_ref:,.2f}** ({ref_source}). "
                "If this keeps happening, the Monday-open capture in "
                "`weekly_setup` for this week is bad — re-run **Weekly Setup** "
                "to overwrite it."
            )
            st.session_state[ref_key] = default_ref
        elif _dev_cur > _REF_WARN_DEV:
            ref_guard_msg = (
                f"⚠️ The {ticker} reference ({float(_cur_ref):,.2f}) is "
                f"{_dev_cur:.0%} away from live spot ({spot:,.2f}). That's an "
                "extreme gap for a weekly anchor — verify the reference before "
                "trusting the strikes below."
            )

    st.html(
        f'<div class="sf-section-title">{ticker} Weekly Credit Spread Finder</div>'
        '<div class="sf-section-sub">HAR regression range forecast · GEX calibration '
        'pending — not applied to live buffer or strikes · 💾 Neon Postgres</div>'
    )
    if ref_guard_msg:
        st.warning(ref_guard_msg)
    if self_heal_msg:
        st.caption(self_heal_msg)

    # Hardcoded macro calendars (FOMC/CPI/NFP/…) lapse yearly; when they do,
    # event weeks silently flag event-free and the buffer stops widening.
    # The log-only warning is invisible on Streamlit Cloud, so surface it.
    try:
        from range_finder.event_calendars import calendar_staleness_warnings
        _cal_warnings = calendar_staleness_warnings()
        if _cal_warnings:
            st.warning("⚠ Event calendar maintenance needed:\n\n"
                       + "\n\n".join(f"• {m}" for m in _cal_warnings))
    except Exception:
        pass

    # ── Extract GEX context from current dashboard data ──
    gex_ctx = extract_gex_context(levels, spot, regime)

    # ── Sidebar-like controls within the tab ──
    col_ctrl1, col_ctrl2, col_ctrl3 = st.columns(3)

    with col_ctrl1:
        step_size = ticker_cfg["strike_increment"]
        # No `value=` — the key is always pre-seeded in session state above,
        # and passing both makes Streamlit log a "widget created with a
        # default value but also had its value set via Session State"
        # warning on every rerun.
        # Wide bounds on purpose: a tight min_value made the frontend clamp
        # any out-of-range value to a plausible-looking bound (e.g. 100 for
        # SPX) instead of letting the sanity guard above see — and heal —
        # the real garbage value.
        spx_close_input = st.number_input(
            f"{ticker} Reference ({ref_source})",
            min_value=1.0, max_value=100000.0, step=float(step_size),
            help=f"Reference price for range calculation. Source: {ref_source}. "
                 "Locked to this week's Monday open Mon-Thu (frozen live Monday, "
                 "restored from the weekly_setup table, or captured retroactively) "
                 "so strikes stay fixed all week; editable if you need to override.",
            key=ref_key,
        )

    with col_ctrl2:
        vix_source = "Mon open" if (frozen_week == current_week and frozen_vix) else "last close"
        vix_input = st.number_input(
            f"VIX Level ({vix_source})",
            min_value=5.0, max_value=100.0, step=0.5,
            help=f"VIX level for BSM credit estimation. Source: {vix_source}. "
                 "Frozen at Monday's open alongside the reference price.",
            key=vix_key,
        )

    with col_ctrl3:
        _spec_keys = list(RF_MODEL_SPECS.keys())
        # Heal a stale per-ticker selection: a session that still has a removed
        # spec (e.g. M6_regime) stored under this key would otherwise crash
        # st.selectbox ("default value not in options"). Drop it so the
        # per-ticker default below applies cleanly.
        _mc_key = f"sf_model_choice_{ticker}"
        if _mc_key in st.session_state and st.session_state[_mc_key] not in _spec_keys:
            del st.session_state[_mc_key]
        _default_model = _default_model_for_ticker(ticker)
        _default_idx = _spec_keys.index(_default_model) if _default_model in _spec_keys else 0
        model_choice = st.selectbox(
            "Model Spec",
            options=_spec_keys,
            index=_default_idx,
            help=(
                "Default is set per ticker from walk-forward OOS: M3_extended for "
                "indices/ETFs and M2_vix for single names. GEX remains excluded "
                "from every live fit while BUG-06 calibration is pending. Your "
                "pick is remembered per ticker."
            ),
            key=f"sf_model_choice_{ticker}",
        )

    # Action buttons
    col_btn_main, col_btn1, col_btn2, col_btn3, col_btn4 = st.columns([1.3, 1, 1, 1, 1])

    with col_btn_main:
        do_weekly = st.button("Weekly Setup", key=f"sf_weekly_{ticker}", type="primary", use_container_width=True,
                              help="Run all steps: Refresh → Rebuild → Save GEX for research → Forecast")
    with col_btn1:
        do_refresh = st.button("Refresh Data", key=f"sf_refresh_{ticker}", use_container_width=True)
    with col_btn2:
        do_rebuild = st.button("Rebuild Features", key=f"sf_rebuild_{ticker}", use_container_width=True)
    with col_btn3:
        do_save_gex = st.button("Save GEX", key=f"sf_save_gex_{ticker}", use_container_width=True)
    with col_btn4:
        do_forecast = st.button("Forecast", key=f"sf_forecast_{ticker}", use_container_width=True)

    # Weekly Setup runs all four steps in sequence
    if do_weekly:
        do_refresh = do_rebuild = do_save_gex = do_forecast = True
        # Weekly Setup may land a fresh weekly_setup row (via the cron
        # path that mirrors Monday's open). Drop the cached lookup so
        # the next render picks it up immediately instead of waiting
        # for the 15 min TTL to expire.
        _cached_weekly_setup.clear()

    conn = _get_rf_conn()
    from phase1.ticker_config import uses_own_har as _uses_own_har
    from phase1.ticker_config import feature_source_ticker as _feature_source

    # ── Step 1: Refresh market data ──
    if do_refresh:
        with st.spinner("1/4 — Fetching SPX / VIX weekly data..."):
            try:
                df_spx = rf_fetch_spx_vix(years=6)
                rows_written = rf_save_spx_vix(conn, df_spx)
                if len(df_spx) == 0:
                    st.success("SPX/VIX data already up to date")
                else:
                    st.success(f"SPX/VIX data refreshed — {len(df_spx)} weeks fetched, {rows_written} new")
            except Exception as e:
                if "empty" in str(e).lower() and datetime.today().weekday() >= 4:
                    st.warning(f"SPX/VIX fetch returned empty data (expected on weekends/holidays). Existing data is still valid.")
                else:
                    st.error(f"SPX/VIX fetch failed: {e}")

        # Own-HAR tickers (QQQ/AMZN/AMD) model their own weekly OHLC +
        # vol proxy, not SPX's. The cron always refreshed both; this
        # button used to refresh only SPX/VIX, so Weekly Setup on those
        # tickers fit against stale weekly_underlying rows.
        if _uses_own_har(ticker):
            with st.spinner(f"1/4 — Fetching {ticker} weekly data..."):
                try:
                    from range_finder.data_collector import (
                        fetch_underlying_weekly, save_underlying_weekly,
                    )
                    df_t = fetch_underlying_weekly(
                        ticker=ticker,
                        yf_symbol=ticker_cfg["yf_symbol"],
                        vol_proxy_yf=ticker_cfg["vol_proxy_yf"],
                        years=6,
                    )
                    _n_t = save_underlying_weekly(conn, ticker, df_t)
                    st.success(f"{ticker} weekly data refreshed — {len(df_t)} weeks fetched, {_n_t} upserted")
                except Exception as e:
                    st.error(f"{ticker} weekly fetch failed: {e}")

        with st.spinner("1/4 — Fetching FRED macro data..."):
            try:
                df_macro = rf_fetch_fred_macro(years=6)
                rf_save_fred_macro(conn, df_macro)
                st.success(f"FRED macro data refreshed — {len(df_macro)} rows")
            except Exception as e:
                # Distinguish "no key" from "FRED returned an error" — the
                # old message lumped them together and blamed the user for
                # a missing key whenever FRED itself had a 500. Also
                # surface the key status so you can eyeball whether
                # Streamlit actually picked up the secret.
                if not RF_FRED_API_KEY:
                    st.warning(
                        "FRED fetch skipped: FRED_API_KEY is not set. "
                        "Add it under Streamlit Cloud → Manage app → Secrets, "
                        "or export FRED_API_KEY in your local env."
                    )
                else:
                    st.warning(
                        f"FRED fetch failed — {e}. "
                        f"Key status: {rf_fred_key_status()}. "
                        "Existing macro data in the DB is still valid; "
                        "the rest of the pipeline will keep running. "
                        "Try again in a minute — FRED's API occasionally "
                        "returns 500s during their maintenance windows."
                    )

        rf_build_event_flags(conn)

    # ── Step 2: Rebuild features ──
    if do_rebuild:
        with st.spinner("2/4 — Computing feature matrix..."):
            try:
                # Same per-ticker routing as the Monday cron: a scaled mini
                # (XSP→SPX) shares its parent's feature rows; own-HAR
                # tickers build their own. (This used to always build SPX, so
                # Rebuild on QQQ/AMZN/AMD never touched the rows the fit reads.)
                rf_build_features(conn, ticker=_feature_source(ticker))
                # Drop the cached feature frame so the next read reflects
                # the rebuild — otherwise the UI would keep serving the
                # pre-rebuild rows until the 10-minute TTL expires.
                _cached_rf_get_features.clear()
                st.success("Features rebuilt")
            except Exception as e:
                st.error(f"Feature rebuild failed: {e}")

    # ── Step 3: Save live GEX ──
    if do_save_gex:
        try:
            gex_flag = save_gex_to_range_finder(gex_ctx, conn, ticker=ticker)
            regime_label = {1: "positive", 0: "neutral", -1: "negative"}.get(gex_flag, "unknown")
            # Invalidate the weekly_setup cache so a follow-up render sees
            # any concurrently-saved row right away (defensive — Save GEX
            # writes gex_inputs not weekly_setup, but the click is the
            # canonical "I'm actively setting up this week" signal).
            _cached_weekly_setup.clear()
            st.success(
                f"GEX research capture saved: regime={regime_label}, "
                f"flag={gex_flag} (ticker={ticker})"
            )
        except Exception as e:
            st.error(f"GEX save failed: {e}")

    st.markdown("---")

    # ── Check data availability (reload after rebuild if needed) ──
    # A scaled mini (XSP→SPX) rides its parent's HAR features at
    # 1/divisor scale (see ticker_config shares_har_with). Own-HAR tickers
    # (QQQ / SPY / NDX / AMZN / AMD) read their own per-ticker feature rows.
    from phase1.ticker_config import feature_source_ticker, get_config as _get_cfg
    _features_ticker = feature_source_ticker(ticker)
    try:
        df_feat = _cached_rf_get_features(conn, ticker=_features_ticker)
    except Exception:
        df_feat = pd.DataFrame()

    if df_feat.empty:
        # Own-HAR tickers cold-start automatically the first time they're
        # looked up: pull history → build
        # features → fit every spec. SPX (the base) keeps the manual nudge — an
        # empty SPX matrix means the whole app is uninitialized, a deploy-time
        # bootstrap step, not a per-ticker warm-up.
        _warm_attempted = st.session_state.setdefault("_sf_warmup_attempted", set())
        if _features_ticker != "SPX" and _features_ticker not in _warm_attempted:
            _warm_attempted.add(_features_ticker)
            with st.spinner(
                f"Preparing spread model for {ticker} (one-time, ~10s) — "
                "pulling history and fitting all specs…"
            ):
                _ok = _auto_warm_up_spread_model(
                    conn, _features_ticker, _get_cfg(_features_ticker)
                )
            if _ok:
                try:
                    df_feat = _cached_rf_get_features(conn, ticker=_features_ticker)
                except Exception:
                    df_feat = pd.DataFrame()

        if df_feat.empty:
            if _features_ticker != "SPX":
                st.warning(
                    f"📉 Spread Finder isn't available for **{ticker}** — no usable "
                    "multi-year price history was found for it (the GEX view above "
                    "still works). This typically means Yahoo Finance has no history "
                    "under the same symbol. Try **Refresh Market Data → Rebuild "
                    "Features → Forecast** to retry."
                )
            else:
                st.info(
                    "No feature data found. Click **Refresh Market Data** then **Rebuild Features** "
                    "to initialize the range prediction model (requires FRED API key in environment)."
                )
            # Still show GEX context even without model data
            _render_gex_context_panel(gex_ctx, spot)
            return

    # ── Fit model or load from cache ──
    # ── Step 4: Fit model and forecast ──
    # Session-state layout: we keep the fit result, features, metrics and
    # the *name* of the spec that produced them.  Tracking the name lets
    # the dropdown actually work — without it, switching M3→M4 would
    # silently keep using the old M3 fit (metric cards and strikes both)
    # because the rest of the code just reads `sf_model_result_{ticker}`
    # regardless of what the selectbox currently says.
    _mdl_result_key  = f"sf_model_result_{ticker}"
    _mdl_feat_key    = f"sf_model_features_{ticker}"
    _mdl_metrics_key = f"sf_model_metrics_{ticker}"
    _mdl_name_key    = f"sf_model_name_{ticker}"

    if do_forecast:
        # Weekly Setup fits every spec (same as the Monday cron) so a
        # user can switch the model dropdown afterwards without tripping
        # the "click Forecast to fit this spec" prompt. A standalone
        # Forecast click still fits only the currently-selected spec —
        # that's the fast path for iterating on one model.
        _specs_to_fit = list(RF_MODEL_SPECS.keys()) if do_weekly else [model_choice]

        # A shared pair (SPX+XSP) reuses one scale-invariant fit, so a
        # Weekly Setup click on either member populates the whole group — every
        # curated ticker that resolves to the same feature source. Own-HAR
        # tickers (QQQ / SPY / AMZN / AMD) stand alone. A plain Forecast click
        # always saves only to the active ticker.
        from phase1.ticker_config import feature_source_ticker as _fsrc, all_tickers as _all_t
        if do_weekly:
            _src = _fsrc(ticker)
            _tickers_to_save = [t for t in _all_t() if _fsrc(t) == _src] or [ticker]
        else:
            _tickers_to_save = [ticker]

        _spinner_label = (
            f"4/4 — Fitting {len(_specs_to_fit)} specs..."
            if len(_specs_to_fit) > 1 else f"4/4 — Fitting {model_choice}..."
        )

        _selected_result = None
        _selected_avail  = None
        _selected_metrics = None

        with st.spinner(_spinner_label):
            for _spec in _specs_to_fit:
                try:
                    # BUG-06 filters GEX from every live fit while preserving
                    # the stored column for the pending historical study.
                    feat_cols = live_spread_feature_columns(RF_MODEL_SPECS[_spec])

                    # Keep the former M4 injection path behind the same policy
                    # switch as run_full_pipeline. The false branch documents
                    # the research retention without consuming the feature.
                    if _spec == "M4_full":
                        gex_col = GEX_NORMALIZED_FEATURE
                        if (
                            GEX_LIVE_SPREAD_INFLUENCE_ENABLED
                            and rf_feature_has_enough_data(df_feat, gex_col)
                        ):
                            if gex_col not in feat_cols:
                                feat_cols.append(gex_col)
                                # Only surface the "using N weeks of GEX" note
                                # when fitting the spec the user is looking at
                                # — otherwise the Weekly Setup run spams the
                                # UI with notes about every background spec.
                                if _spec == model_choice:
                                    st.caption(
                                        f"ℹ️ M4_full: using {int(df_feat[gex_col].notna().sum())} "
                                        f"weeks of stored GEX history as a training feature."
                                    )
                        elif GEX_LIVE_SPREAD_INFLUENCE_ENABLED:
                            _weeks = int(df_feat[gex_col].notna().sum()) if gex_col in df_feat.columns else 0
                            if _spec == model_choice:
                                st.caption(
                                    f"ℹ️ M4_full: only {_weeks} weeks of GEX history — need >{RF_GEX_MIN_WEEKS} to "
                                    f"fold `gex_normalized` into the fit. Keep clicking **Save GEX** "
                                    f"each week; in the meantime M4 runs without the GEX feature."
                                )
                        elif _spec == model_choice:
                            _weeks = (
                                int(df_feat[gex_col].notna().sum())
                                if gex_col in df_feat.columns else 0
                            )
                            st.caption(
                                f"GEX calibration: Pending — {_weeks} stored weeks "
                                "remain available for research, but GEX is not "
                                "applied to live fits, buffers, or strikes."
                            )

                    avail_cols = production_feature_columns(df_feat, _spec)
                    if len(avail_cols) < 2:
                        if _spec == model_choice:
                            st.warning(f"{_spec}: only {len(avail_cols)} usable features — skipped")
                        continue

                    _validation, _result, _metrics = (
                        rf_fit_validation_and_production(
                            df_feat, feature_cols=avail_cols, model_name=_spec
                        )
                    )
                    for _save_ticker in _tickers_to_save:
                        rf_save_model(_result, avail_cols, _spec, _metrics,
                                      conn=conn, ticker=_save_ticker)
                    # New fit landed in Postgres — drop any cached
                    # unpickle for this (spec, ticker) pair so a later
                    # tab switch sees the fresh weights instead of
                    # serving the previous fit until the 1-hour TTL
                    # expires.
                    _cached_rf_load_model.clear()

                    if _spec == model_choice:
                        _selected_result  = _result
                        _selected_avail   = avail_cols
                        _selected_metrics = _metrics
                except Exception as e:
                    st.error(f"{_spec} fitting failed: {e}")

        # Summary line for the Weekly Setup path so the user can see which
        # specs landed in Postgres at a glance.
        if do_weekly:
            _ticker_note = " × ".join(_tickers_to_save)
            st.success(
                f"Fitted {len(_specs_to_fit)} specs for {_ticker_note} "
                f"({len(_specs_to_fit) * len(_tickers_to_save)} rows saved)"
            )

        # Prime session state with the currently-selected spec's fit so
        # the rest of this render uses it without falling through to the
        # load-from-Postgres path (same behavior as the previous single-
        # spec code).
        if _selected_result is not None:
            st.session_state[_mdl_result_key]  = _selected_result
            st.session_state[_mdl_feat_key]    = _selected_avail
            st.session_state[_mdl_metrics_key] = _selected_metrics
            st.session_state[_mdl_name_key]    = model_choice
            st.success(
                f"Model fitted | {model_choice} | "
                f"OOS R² = {_selected_metrics['oos_r2']:.4f}"
            )

    # If the user toggled the model dropdown, the cached fit in session
    # state belongs to a different spec — evict it so the load block
    # below pulls the right saved fit for the newly-selected spec from
    # Postgres (or shows the "click Forecast" nudge if that spec has
    # never been fitted yet).
    _cached_mdl_name = st.session_state.get(_mdl_name_key)
    if _cached_mdl_name is not None and _cached_mdl_name != model_choice:
        for _k in (_mdl_result_key, _mdl_feat_key, _mdl_metrics_key, _mdl_name_key):
            st.session_state.pop(_k, None)

    # Try to load model from session or disk
    if _mdl_result_key not in st.session_state:
        from phase1.ticker_config import feature_source_ticker as _fsrc
        try:
            try:
                payload = _cached_rf_load_model(model_choice, ticker)
            except FileNotFoundError:
                # A scaled mini (XSP→SPX) rides its parent's fit
                # (identical by construction) when it has no row of its own.
                _parent = _fsrc(ticker)
                if _parent == ticker:
                    raise
                payload = _cached_rf_load_model(model_choice, _parent)
            st.session_state[_mdl_result_key]  = payload["result"]
            st.session_state[_mdl_feat_key]    = payload["feature_cols"]
            st.session_state[_mdl_metrics_key] = payload["metrics"]
            st.session_state[_mdl_name_key]    = model_choice
        except FileNotFoundError:
            st.info(f"No saved fit for **{model_choice}** yet. Click **Forecast** to fit it for the first time.")
            _render_gex_context_panel(gex_ctx, spot)
            return
        except Exception as e:
            st.warning(f"Saved {model_choice} model incompatible: {e}. Click **Forecast** to refit.")
            _render_gex_context_panel(gex_ctx, spot)
            return

    result       = st.session_state[_mdl_result_key]
    feat_cols    = st.session_state[_mdl_feat_key]
    metrics      = st.session_state[_mdl_metrics_key]
    active_model = st.session_state.get(_mdl_name_key, model_choice)

    if uses_disabled_gex_feature(feat_cols):
        for _k in (_mdl_result_key, _mdl_feat_key, _mdl_metrics_key, _mdl_name_key):
            st.session_state.pop(_k, None)
        st.warning(
            f"Saved {model_choice} fit predates the BUG-06 mitigation and "
            "contains GEX. Click Forecast to refit it without GEX; no live "
            "recommendation will be shown from the old fit."
        )
        _render_gex_context_panel(gex_ctx, spot)
        return

    # ── Determine week start ──
    # Anchored to NY wall clock (run_now = now_ny() above) so that the
    # week_start convention here matches the Friday chain pre-fetch in
    # streamlit_app.fetch_all_data — otherwise a UTC-hosted server could
    # roll into "tomorrow" a few hours before NY does and end up looking
    # at a different expiration than the one the pre-fetch cached.
    #
    # Mon-Thu: plan THIS week's spreads (week_start = this Monday, expiring
    # this Friday).  Fri-Sun: this week is done, so roll forward to next
    # Monday's week.  This has to match _spread_finder_target_friday above.
    _wd = run_now.weekday()
    if _wd <= 3:                           # Mon-Thu
        monday_dt = run_now - timedelta(days=_wd)
    else:                                  # Fri-Sun
        monday_dt = run_now + timedelta(days=(7 - _wd))
    week_start = monday_dt.strftime("%Y-%m-%d")
    sf_ref_date = run_now.date()

    # ── Get feature row ──
    # Look up the row in the already-cached `df_feat` instead of issuing a
    # fresh `SELECT * FROM model_features WHERE week_start = ?` on every
    # render. `df_feat` is loaded by `_cached_rf_get_features` (10 min TTL)
    # and indexed by `week_start` (DatetimeIndex), so this is a pure
    # in-memory lookup and saves one Neon roundtrip per Spread Finder pass.
    feature_row = None
    try:
        _wk_ts = pd.Timestamp(week_start)
        if _wk_ts in df_feat.index:
            feature_row = df_feat.loc[_wk_ts]
    except Exception:
        feature_row = None
    if feature_row is None:
        _fallback_idx = df_feat.index[-1]
        _fallback_label = (
            _fallback_idx.strftime("%Y-%m-%d")
            if hasattr(_fallback_idx, "strftime") else str(_fallback_idx)
        )
        st.error(
            f"⚠️ **Forecast blocked:** no feature row exists for {week_start}. "
            f"The newest persisted row is {_fallback_label}. Run **Weekly "
            "Setup** (or **Refresh Data** + **Rebuild Features**) before "
            "using this week's strikes."
        )
        _render_gex_context_panel(gex_ctx, spot)
        return

    _path_fresh, _path_source, _path_required = rf_assess_path_provenance(
        feature_row, _wk_ts,
    )
    if not _path_fresh:
        _source_label = (
            _path_source.strftime("%Y-%m-%d")
            if _path_source is not None else "missing"
        )
        st.error(
            f"⚠️ **Forecast blocked:** the {week_start} row uses weekly path "
            f"inputs from **{_source_label}**, but it requires "
            f"**{_path_required:%Y-%m-%d}**. A provisional next-week row "
            "created before Friday's close cannot be reused after the week "
            "rolls. Run **Weekly Setup** (or **Refresh Data** + **Rebuild "
            "Features**) to refresh the HAR inputs. The saved Monday-open "
            "strike anchor will remain unchanged."
        )
        _render_gex_context_panel(gex_ctx, spot)
        return

    # ── Regime-shift circuit breaker ──────────────────────────────────────
    # HAR features are lagged by one week — vix_close in feature_row is
    # last Friday's close. When IV spikes overnight (e.g., VIX 15 → 40 on
    # a news shock), the model's input features still reflect the pre-spike
    # world for a full week, so the forecast's PI is anchored to the wrong
    # vol regime and the Spread Finder will place strikes dangerously
    # close to spot. The live weekly expected move (from the straddle) is
    # drawn on the strike map as a reference band and shorts inside it are
    # flagged, but that doesn't stop the user from trusting the "model says
    # range will be 2%" read.
    #
    # Detect the shift by comparing the live VIX to the trailing VIX
    # already in the feature row. Ratio > 1.5 is the threshold — that's
    # roughly a 2σ move on the weekly VIX change distribution.
    #
    # Use live_vix_raw (the genuine fetched value), NOT live_vix — the latter
    # falls back to a fabricated 18.0 when the fetch fails, and comparing a
    # made-up number against the trailing VIX would either miss a real spike
    # or invent a phantom one. If we have no live VIX, we simply can't judge
    # a regime shift, so the breaker stays silent.
    _trailing_vix = None
    try:
        _trailing_vix_raw = feature_row.get("vix_close")
        if _trailing_vix_raw is not None:
            _trailing_vix = float(_trailing_vix_raw)
    except Exception:
        pass

    regime_shift = None
    if _trailing_vix and _trailing_vix > 0 and live_vix_raw and live_vix_raw > 0:
        _vix_ratio = live_vix_raw / _trailing_vix
        if _vix_ratio >= 1.5:
            regime_shift = {
                "severity": "extreme" if _vix_ratio >= 2.0 else "elevated",
                "live_vix": live_vix_raw,
                "trailing_vix": _trailing_vix,
                "ratio": _vix_ratio,
            }

    if regime_shift is not None:
        _sev_word = regime_shift["severity"]
        st.error(
            f"⚠️ **VIX regime shift detected ({_sev_word})** — "
            f"live VIX **{regime_shift['live_vix']:.1f}** vs trailing "
            f"feature VIX **{regime_shift['trailing_vix']:.1f}** "
            f"(**{regime_shift['ratio']:.2f}×**).\n\n"
            f"The HAR model's features lag by one week, so the forecast below "
            f"is anchored to the pre-spike vol regime. Short strikes sized "
            f"against this forecast are likely **too narrow**. Check the live "
            f"weekly expected-move band on the strike map (it reflects the "
            f"current straddle) and treat any short strike inside it as too "
            f"risky — or, better, skip the trade until features catch up."
        )
    # ── Build forecast → plan → tiers from the latest GEX refresh ──
    # We intentionally DO NOT cache these on a session-state key any more.
    # Everything below is cheap arithmetic on top of the already-loaded HAR
    # model (the expensive validation + production fits are gated behind the
    # "Forecast" button and cached separately via rf_load_model), so
    # recomputing on every page rerun lets the spread finder pick up fresh
    # chain bid/ask as soon as fetch_all_data refreshes data.chain_cache
    # (i.e. on auto-refresh, "Refresh Now", or any normal rerun — no need
    # to click "Forecast" again to get updated credits).
    #
    # Risk-tier switching stays snappy because the _risk_tier_fragment
    # below is wrapped in @st.fragment and only re-reads the spread_tiers
    # we stash in session_state — the outer recompute doesn't happen on
    # tier toggles.
    # Per-side band share: empirical side-share quantile (falls back to the
    # legacy /2 split when weekly history is unavailable). Cached 1h.
    side_share_q = _cached_side_share_q(ticker)

    chain_quotes, chain_exp = _build_chain_quotes_for_spreads(
        data, ticker, ref_date=sf_ref_date,
    )
    forecast, plan, spread_tiers = build_recommendations(
        result=result, feature_row=feature_row, feature_cols=feat_cols,
        reference=spx_close_input, vix=vix_input, week_start=week_start,
        ticker=ticker, side_share_q=side_share_q, chain_quotes=chain_quotes,
        weekly_em=weekly_em, conn=conn, model_name=model_choice,
    )

    gex_adj = adjust_spread_with_gex(plan, gex_ctx)

    # =========================================================================
    # METRIC CARDS
    # =========================================================================

    # Five metric cards. The PI card holds both tier bounds (Lower ↔ Upper)
    # so all four risk tiers named in the spec — Lower PI, Point, Upper PI,
    # Effective — stay visible without cramping the row into 6 columns
    # (which clips labels on typical laptop widths).
    # Flag it when the dropdown's selection and the fit we're actually
    # rendering don't agree (the OOS card should tell the truth, not lie).
    _mdl_label = active_model if active_model == model_choice else f"{active_model} ⚠"
    _flag = gex_adj['gex_regime_flag']
    _reg_color = ("var(--green)" if _flag > 0 else
                  "var(--red)" if _flag < 0 else "var(--text-muted)")

    def _mc(label, value, sub, value_color="var(--text-primary)"):
        return (f'<div class="sf-card"><div class="cl">{label}</div>'
                f'<div class="cv" style="color:{value_color};">{value}</div>'
                f'<div class="cs">{sub}</div></div>')

    st.html(
        '<div class="sf-cards">'
        + _mc("Point Estimate", f"{forecast['point_pct']*100:.2f}%",
              f"vs VIX {forecast['model_vs_vix']*100:+.2f}%")
        + _mc(f"{forecast['confidence_level']}% PI Range",
              f"{forecast['lower_pct']*100:.2f}–{forecast['upper_pct']*100:.2f}%",
              "aggressive ↔ moderate")
        + _mc("Effective Range", f"{plan.effective_range_pct*100:.2f}%",
              f"conservative · buffer +{plan.buffer_pct*100:.2f}%")
        + _mc("GEX Regime", gex_ctx.gamma_regime.title(),
              "research · not applied", _reg_color)
        + _mc(f"OOS R² · {_mdl_label}", f"{metrics['oos_r2']:.4f}",
              f"MAE {metrics['mae_pct']*100:.2f}%"
              + (f" · dir {metrics['direction_acc']:.0%}"
                 if metrics.get("direction_acc") is not None else ""))
        + '</div>'
    )

    # ── Empirical PI-coverage caption (the calibration audit, in one line) ──
    # Reads scored spread_log rows via calibration.weekly_pi_coverage; the
    # Monday cron logs a plan and scores last week's, so this accumulates one
    # observation per week. Cached so reruns don't hit Neon.
    try:
        _cov = _cached_pi_coverage(ticker)
        if _cov["n"] == 0:
            _cov_txt = "PI calibration: no scored weeks yet — accumulating"
        elif not _cov["sufficient"]:
            _cov_txt = (f"PI calibration: {_cov['one_sided']:.0%} of {_cov['n']} "
                        f"scored weeks inside the PI upper — accumulating "
                        f"(needs 15+ for conclusions)")
        else:
            _buf = _cov.get("buffer", {})
            _buf_txt = (f" · buffer breach {_buf['breach_rate']:.0%}"
                        if _buf.get("n") else "")
            _cov_txt = (f"PI calibration: {_cov['one_sided']:.0%} one-sided over "
                        f"{_cov['n']} scored weeks (nominal "
                        f"{_cov['nominal_one_sided']:.0%}){_buf_txt}")
        st.caption(_cov_txt)
    except Exception:
        pass   # calibration display is never worth breaking the finder over

    # ── Excel export — ALL tickers, one week-named tab ──
    # The active ticker's row reuses the exact tiers rendered above (chain-
    # snapped strikes, live weekly-EM floor); the other four instruments are
    # rebuilt from persisted state (saved fits, Monday-open captures, DB EM
    # snapshots) via a 10-min cross-session cache so the eager download-button
    # payload build doesn't hammer Neon on every rerun.
    # ── Managed export-ticker list ──────────────────────────────────
    # Always-present defaults + whatever the user has added. The active ticker
    # is included only when it's a default or has been explicitly added, so
    # nothing gets silently swept in/out when switching tickers.
    _extras = [t for t in _xlsx_extra_list() if t not in _FT_DEFAULT_TICKERS]
    _export_tickers = _FT_DEFAULT_TICKERS + _extras
    if len(_export_tickers) > _FT_MAX_TICKERS:
        # The Scoreboard scans a fixed 40-row window per week tab, so the
        # workbook builder truncates silently — surface the drop here instead.
        st.caption(
            f"⚠ The export caps at {_FT_MAX_TICKERS} tickers — dropped: "
            f"{', '.join(_export_tickers[_FT_MAX_TICKERS:])}. Remove some "
            "added tickers to include them."
        )

    st.html('<div class="sf-eyebrow">Tickers in this export</div>')
    st.markdown(
        '<div style="font-size:11px;color:var(--text-muted);margin-bottom:6px;">'
        'Always included: <span style="font-family:var(--mono);color:var(--text-secondary);">'
        + ' · '.join(_FT_DEFAULT_TICKERS)
        + '</span></div>',
        unsafe_allow_html=True,
    )
    # Added tickers as small removable chips — click a chip (✕) to drop it.
    if _extras:
        st.markdown(
            '<div style="font-size:9.5px;color:var(--text-dim);text-transform:uppercase;'
            'letter-spacing:.04em;margin:2px 0;">Added (click ✕ to remove)</div>',
            unsafe_allow_html=True,
        )
        st.pills(
            "Added export tickers",
            [f"{t}  ✕" for t in _extras],
            selection_mode="single",
            key="_sf_xlsx_rm_pills",
            label_visibility="collapsed",
            on_change=_cb_remove_extra_pill,
        )
    # Add the active ticker — compact button, only when it isn't already in.
    if ticker not in _export_tickers:
        st.button(
            f"➕ Add {ticker} to export",
            key=f"_sf_xlsx_add_{ticker}",
            on_click=_xlsx_add_extra, args=(ticker,),
        )
    elif ticker in _extras:
        st.caption(f"✓ {ticker} is in the export")

    _xlsx_col, _ = st.columns([1, 3])
    with _xlsx_col:
        try:
            _active_notes = [active_model]
            if gex_ctx.gamma_regime:
                _gflag = gex_adj.get("gex_regime_flag")
                _active_notes.append(
                    f"GEX research {gex_ctx.gamma_regime}"
                    + (f" ({_gflag:+d})" if isinstance(_gflag, int) else "")
                    + " — not applied"
                )
            _active_events = [n for n, f in (("FOMC", plan.has_fomc), ("CPI", plan.has_cpi),
                                             ("NFP", plan.has_nfp), ("OPEX", plan.has_opex)) if f]
            if _active_events:
                _active_notes.append("/".join(_active_events))
            if plan.buffer_pct:
                _active_notes.append(f"buf {plan.buffer_pct * 100:.2f}%")
            if plan.recommended_width:
                _active_notes.append(f"wing {plan.recommended_width:g}")
            if chain_exp:
                _active_notes.append(f"chain {chain_exp}")

            _active_row = {
                "ticker": ticker,
                "ref": round(float(spx_close_input), 2),
                "prev_close": _cached_prior_week_close(week_start, ticker),
                "bands": _tier_bands_from_tiers(spread_tiers),
                "notes": _active_notes,
                "error": None,
            }
            # Rebuild every NON-active export ticker from persisted state; slot
            # the live active row in if the active ticker is in the export list.
            _other_tickers = tuple(t for t in _export_tickers if t != ticker)
            _other_rows = {
                r["ticker"]: r
                for r in _cached_nonactive_week_bands(week_start, active_model, _other_tickers)
            }
            _ft_rows = [
                _active_row if t == ticker else _other_rows.get(
                    t, {"ticker": t, "error": "data collection failed",
                        "bands": {}, "notes": [], "ref": None, "prev_close": None},
                )
                for t in _export_tickers
            ]

            # Memoize the workbook on a content signature so an unrelated
            # widget rerun (ticker chip toggle, tier switch, auto-refresh
            # tick) doesn't rebuild the whole openpyxl workbook + formula
            # grid every pass — only an actual change to the rows/bands does.
            import hashlib as _hl, json as _json
            _sig = _hl.sha1(
                _json.dumps(_ft_rows, sort_keys=True, default=str).encode()
            ).hexdigest()
            _xlsx_sig_key = f"_sf_xlsx_sig_{week_start}_{active_model}"
            _xlsx_bytes_key = f"_sf_xlsx_bytes_{week_start}_{active_model}"
            if st.session_state.get(_xlsx_sig_key) != _sig:
                st.session_state[_xlsx_bytes_key] = _build_forward_test_workbook(
                    week_start=week_start, model_choice=active_model, rows=_ft_rows,
                )
                st.session_state[_xlsx_sig_key] = _sig
            _xlsx_bytes = st.session_state[_xlsx_bytes_key]
            st.download_button(
                label       = "Export weekly workbook",
                data        = _xlsx_bytes,
                file_name   = f"har_forward_test_{week_start}_{active_model}.xlsx",
                mime        = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key         = f"_sf_xlsx_{week_start}_{active_model}",
                help        = (
                    "One tab named for this week's Monday with each instrument's "
                    "POINT / PI / EFFECTIVE bands, plus a ticker-keyed Scoreboard: "
                    "week tabs are matched by Instrument name, so the ticker mix "
                    "and row order can differ week to week. First download = your "
                    "master workbook; each later week, Move/Copy the new week tab "
                    "into it (any position). Re-importing a week already in the "
                    "master? Delete its old tab first — Excel silently renames "
                    "the paste to '(2)', which the Scoreboard ignores. Fill "
                    "'Weekly Close' after Friday and everything scores itself. "
                    "Upgrading an old master: delete its Scoreboard and bookend "
                    "tabs, Move/Copy this file's Scoreboard + FTBlank sheets in "
                    "(Ctrl+click both), and re-type any old tickers into blank "
                    "slots — history re-scores retroactively."
                ),
            )
            _missing = [r["ticker"] for r in _ft_rows if r.get("error")]
            if _missing:
                st.caption(
                    f"⚠ No bands for {', '.join(_missing)} — see the Notes column "
                    "in the export. Usually fixed by running Weekly Setup on that "
                    "ticker (needs a saved fit + Monday-open capture); the new "
                    "defaults populate after their first cron run."
                )
        except Exception as _xlsx_err:
            st.caption(f"Excel export unavailable: {_xlsx_err}")

    st.markdown("---")

    # =========================================================================
    # RISK TIER SELECTOR + DEPENDENT UI (wrapped in fragment for fast switching)
    # =========================================================================

    # Store everything the fragment needs in session_state so it doesn't
    # rely on closure variables that become stale across fragment reruns.
    st.session_state["_rtf_spread_tiers"] = spread_tiers
    st.session_state["_rtf_forecast"]     = forecast
    st.session_state["_rtf_plan"]         = plan
    st.session_state["_rtf_spx_close"]    = spx_close_input
    st.session_state["_rtf_gex_ctx"]      = gex_ctx
    st.session_state["_rtf_ticker"]       = ticker
    st.session_state["_rtf_weekly_em"]    = weekly_em
    st.session_state["_rtf_chain_exp"]    = chain_exp
    st.session_state["_rtf_spot"]         = spot
    st.session_state["_rtf_gex_adj"]      = gex_adj

    # Previously wrapped in @st.fragment to keep risk-tier clicks from
    # recomputing the plan. That saved ~100ms per tier toggle but the
    # cost was a nested-fragment pattern: outer @st.fragment on
    # _render_spread_finder_tab plus inner @st.fragment on
    # _risk_tier_fragment. Nested fragments schedule reruns
    # independently, and when the inner read `_rtf_*` session-state
    # that the outer had written, rapid ticker/tab switching could
    # leave the inner rendering against a stale snapshot — strikes
    # and widths would stop updating until a hard refresh.  Flattened
    # to a plain nested function: tier toggles now rerun the outer
    # fragment only (still tab-isolated, no full-app rerun), which
    # reuses the cached model fit and is fast enough (~100ms).
    def _risk_tier_fragment():
        _TIER_COLORS = {
            "aggressive":   "#ff4b4b",
            "moderate":     "#ffa726",
            "conservative": "#66bb6a",
        }

        # Read from session_state to avoid stale closure references
        _spread_tiers  = st.session_state["_rtf_spread_tiers"]
        _forecast      = st.session_state["_rtf_forecast"]
        _plan          = st.session_state["_rtf_plan"]
        _spx_close_inp = st.session_state["_rtf_spx_close"]
        _gex_ctx       = st.session_state["_rtf_gex_ctx"]
        _ticker        = st.session_state["_rtf_ticker"]
        _weekly_em     = st.session_state["_rtf_weekly_em"]
        _chain_exp     = st.session_state["_rtf_chain_exp"]
        _spot          = st.session_state["_rtf_spot"]
        _gex_adj       = st.session_state["_rtf_gex_adj"]

        tier_labels = [t.label for t in _spread_tiers]
        default_idx = len(tier_labels) - 1
        st.html('<div class="sf-eyebrow">Risk Tier</div>')
        # One full-width button per tier (st.columns → equal widths, no blank
        # gap), like the action buttons above. The active tier is the primary
        # (green) button; selection persists in session_state.
        _tier_state_key = f"sf_risk_tier_idx_{_ticker}"
        if _tier_state_key not in st.session_state:
            st.session_state[_tier_state_key] = default_idx
        for _i, _col in enumerate(st.columns(len(tier_labels))):
            with _col:
                st.button(
                    f"{tier_labels[_i]} ({_spread_tiers[_i].range_pct*100:.1f}%)",
                    key=f"sf_tier_btn_{_ticker}_{_i}",
                    type=("primary" if _i == st.session_state[_tier_state_key] else "secondary"),
                    use_container_width=True,
                    on_click=lambda idx=_i: st.session_state.__setitem__(_tier_state_key, idx),
                )
        selected_tier_idx = st.session_state[_tier_state_key]

        selected_tier = _spread_tiers[selected_tier_idx]

        # Present the HAR model's ORIGINAL strikes (before any weekly-EM floor).
        # The tier carries both: `.call_short`/`.put_short` are the EM-floored
        # (widened) strikes; `.model_call_short`/`.model_put_short` hold the
        # pre-floor strikes, but ONLY when the floor actually moved them (None
        # otherwise). Per design we no longer snap the short strike out to the
        # EM boundary — we show the model's own strike and flag it when it sits
        # inside the weekly expected move (warning below + EM band on the map).
        # Falling back to the floored value when the model field is None is
        # exact: None means the floor never moved that strike.
        displayed = displayed_tier(selected_tier)
        disp_call_short = displayed["call_short"]
        disp_put_short = displayed["put_short"]
        disp_call_spreads = displayed["call_spreads"]
        disp_put_spreads = displayed["put_spreads"]

        tier_color = _TIER_COLORS.get(selected_tier.risk_level, "#888")
        st.markdown(
            f"<span style='color:{tier_color};font-size:18px;font-weight:bold;'>"
            f"{selected_tier.risk_level.upper()}</span>"
            f" &nbsp;—&nbsp; Range: {selected_tier.range_pct*100:.2f}%"
            f" &nbsp;|&nbsp; Calls above `{_fmt_strike(disp_call_short)}`"
            f" &nbsp;|&nbsp; Puts below `{_fmt_strike(disp_put_short)}`",
            unsafe_allow_html=True,
        )

        # =====================================================================
        # RANGE GAUGE + STRIKE MAP (side by side)
        # =====================================================================

        col_gauge, col_strikes = st.columns([1, 1])

        with col_gauge:
            st.markdown("**Forecast Range**")
            _render_sf_range_gauge(
                _forecast, _plan, _spx_close_inp,
                tier_label=selected_tier.label,
                tier_range_pct=selected_tier.range_pct,
                tier_risk_level=selected_tier.risk_level,
            )

        with col_strikes:
            st.markdown("**Strike Map with GEX Walls**")
            _render_sf_strike_map_tier(
                selected_tier, _plan, _spx_close_inp, _gex_ctx,
                _plan.recommended_width, ticker=_ticker,
                weekly_em=_weekly_em,
            )

        st.markdown("---")

        st.markdown(f"**Spread Parameters — {selected_tier.label}**")

        col_call, col_put = st.columns(2)

        with col_call:
            st.markdown(f"Call Spreads — short above `{_fmt_strike(disp_call_short)}`")
            _render_sf_spread_table(disp_call_spreads)

        with col_put:
            st.markdown(f"Put Spreads — short below `{_fmt_strike(disp_put_short)}`")
            _render_sf_spread_table(disp_put_spreads)

        # Show credit source note with chain expiration
        all_tier_spreads = disp_call_spreads + disp_put_spreads
        has_market = any(getattr(s, "credit_source", "bsm") == "market" for s in all_tier_spreads)
        has_bsm = any(getattr(s, "credit_source", "bsm") == "bsm" for s in all_tier_spreads)
        exp_note = f" Chain: {_chain_exp}" if _chain_exp else ""
        if has_market and has_bsm:
            st.caption(f"Credits from Friday chain bid/ask.{exp_note} &nbsp;|&nbsp; * = BSM estimate (strike not in chain).")
        elif has_market:
            st.caption(f"Credits from Friday chain bid/ask (short bid - long ask).{exp_note}")
        else:
            st.caption("Credits are BSM estimates (no Friday chain data available). Verify with broker before trading.")

        # =====================================================================
        # GEX CONTEXT + WARNINGS
        # =====================================================================

        st.markdown("---")

        col_gex, col_warn = st.columns([1, 1])

        with col_gex:
            _render_gex_context_panel(_gex_ctx, _spot)

        with col_warn:
            st.html('<div class="sf-eyebrow">Warnings &amp; GEX Research Notes</div>')

            # One coherent GEX story: strips the anchored-regime warning and
            # composes a single anchored-vs-live message (flip-aware) instead
            # of concatenating two independent — and possibly contradictory —
            # regime signals.
            all_warnings = reconcile_gex_warnings(
                list(_plan.warnings),
                _gex_adj.get("gex_adjustment_notes", []),
                anchored_gex_flag=_plan.gex_flag,
                live_regime=_gex_adj.get("gex_regime"),
                anchor_label="Monday anchor",
            )

            # Inside-the-expected-move flags. The model no longer snaps the
            # short strike out to the EM boundary (that "EM floor" feature was
            # removed); instead we keep the model's own strike and warn when it
            # sits inside the weekly expected move. The model marks that case by
            # populating `model_*_short` (only set when the original strike was
            # inside the EM), so it's the authoritative inside-EM signal.
            _em_upper = (_weekly_em or {}).get("upper_level", 0) or 0
            _em_lower = (_weekly_em or {}).get("lower_level", 0) or 0
            if _em_upper > 0 and _em_lower > 0:
                if selected_tier.model_call_short is not None:
                    all_warnings.append(
                        f"Call short {_fmt_strike(disp_call_short)} is within the weekly "
                        f"expected move (EM upper {_em_upper:,.0f})."
                    )
                if selected_tier.model_put_short is not None:
                    all_warnings.append(
                        f"Put short {_fmt_strike(disp_put_short)} is within the weekly "
                        f"expected move (EM lower {_em_lower:,.0f})."
                    )

            _esc = lambda s: str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            if all_warnings:
                st.html("".join(
                    f'<div class="sf-note warn" style="margin-bottom:6px;">{_esc(w)}</div>'
                    for w in all_warnings
                ))
            else:
                st.html('<div class="sf-note ok">No warnings for this week.</div>')

            # Event flags
            events = {"FOMC": _plan.has_fomc, "CPI": _plan.has_cpi, "NFP": _plan.has_nfp, "OPEX": _plan.has_opex}
            active = [k for k, v in events.items() if v]
            if active:
                st.html('<div class="sf-note info" style="margin-top:6px;">Events this week: '
                        f'<b style="color:var(--text-secondary);">{", ".join(active)}</b></div>')
            else:
                st.html('<div class="sf-note" style="margin-top:6px;">No major events this week</div>')

    _risk_tier_fragment()


def _render_gex_context_panel(gex_ctx: GEXContext, spot: float):
    """Render live GEX as research context, never a spread input."""
    gex_flag = regime_to_gex_flag(gex_ctx.gamma_regime)
    regime_color = SF_BULL if gex_flag == 1 else SF_BEAR if gex_flag == -1 else SF_NEUT
    zg_pct = (abs(spot - gex_ctx.zero_gamma) / spot * 100) if spot else 0.0

    st.html(
        '<div class="sf-gex">'
        '<div class="sf-eyebrow">Live GEX Context</div>'
        f'<div class="reg" style="color:{regime_color};">{gex_ctx.gamma_regime.title()}</div>'
        '<div class="row"><span class="k">Zero-Gamma</span>'
        f'<span class="v" style="color:var(--cyan);">${gex_ctx.zero_gamma:,.0f} '
        f'<span class="sub">({zg_pct:.2f}% from spot)</span></span></div>'
        '<div class="row"><span class="k">Call Wall</span>'
        f'<span class="v" style="color:var(--green);">${gex_ctx.call_wall:,.0f}</span></div>'
        '<div class="row"><span class="k">Put Wall</span>'
        f'<span class="v" style="color:var(--red);">${gex_ctx.put_wall:,.0f}</span></div>'
        '<div class="row"><span class="k">Spread Influence</span>'
        '<span class="v" style="color:var(--text-secondary);">Research / disabled</span></div>'
        '<div class="row"><span class="k">Spot</span>'
        f'<span class="v">${spot:,.2f}</span></div>'
        '</div>'
    )


def _render_sf_range_gauge(
    forecast: dict,
    plan: SpreadPlan,
    spx_ref: float,
    tier_label: str = None,
    tier_range_pct: float = None,
    tier_risk_level: str = None,
):
    """Horizontal number-line gauge of the weekly range forecast.

    The old view stacked four ascending bars (Lower PI, Point, Upper PI,
    Effective) which read like four independent predictions climbing in
    magnitude. They aren't — they're all positions on the same weekly
    range-% axis: Point is the model's central forecast, Lower/Upper PI
    are the 10th/90th percentiles of its predictive distribution, and
    Effective is Upper PI plus a fixed safety buffer. This view lays
    them out on one horizontal axis:

      - Amber band  = 80% prediction interval (Lower PI → Upper PI)
      - Red band    = buffer extension (Upper PI → Effective)
      - Bull dot    = Point Estimate (the center of the distribution)
      - Dashed tick = VIX-implied range for reference
      - White caret = the Risk Tier currently driving strike placement

    Because the forecast is fit on log(range), the back-transformed
    distribution is asymmetric — Point will usually sit below the
    midpoint of the PI band. That's statistically correct, not a bug,
    and the horizontal layout actually makes it visible (whereas the
    old four-bar chart hid it entirely).
    """
    lower_pct     = forecast["lower_pct"]         * 100
    point_pct     = forecast["point_pct"]         * 100
    upper_pct     = forecast["upper_pct"]         * 100
    effective_pct = plan.effective_range_pct      * 100
    vix_pct       = forecast["vix_implied_pct"]   * 100
    confidence    = forecast["confidence_level"]

    # Axis extends a touch past the biggest value so the right-most marker
    # doesn't collide with the track edge.
    axis_max = max(effective_pct, vix_pct, point_pct) * 1.18 + 0.4
    if axis_max <= 0:
        axis_max = 1.0

    def _p(x: float) -> float:  # value → % position on the axis, clamped
        return max(0.0, min(100.0, x / axis_max * 100.0))

    tier_colors = {"aggressive": "var(--red)", "moderate": "var(--amber)",
                   "conservative": "var(--green)"}
    tier_color = tier_colors.get((tier_risk_level or "").lower(), "#fff")

    lo, up, eff, pt, vx = (_p(lower_pct), _p(upper_pct), _p(effective_pct),
                           _p(point_pct), _p(vix_pct))

    caret = ""
    if tier_range_pct is not None:
        caret = (f'<div class="caret" style="left:{_p(tier_range_pct * 100):.2f}%;'
                 f'color:{tier_color};font-weight:700;">▼ {tier_label or "Selected"}</div>')

    st.html(
        '<div class="sf-gauge">'
        '<div class="track">'
        f'<div class="band" style="left:{lo:.2f}%;width:{max(0.0, up - lo):.2f}%;'
        'background:rgba(255,180,84,.55);border-radius:3px;"></div>'
        f'<div class="band" style="left:{up:.2f}%;width:{max(0.0, eff - up):.2f}%;'
        'background:rgba(255,77,104,.5);"></div>'
        f'<div class="vix" style="left:{vx:.2f}%;"></div>'
        f'<div class="dot" style="left:{pt:.2f}%;"></div>'
        f'{caret}'
        '</div>'
        '<div class="axis"><span>0%</span>'
        f'<span>{axis_max / 2:.1f}%</span><span>{axis_max:.1f}%</span></div>'
        '<div style="display:flex;flex-wrap:wrap;gap:10px 14px;margin-top:10px;'
        'font-family:var(--mono);font-size:10px;">'
        f'<span style="color:var(--amber);">Lower PI {lower_pct:.2f}%</span>'
        f'<span style="color:var(--green);">Point {point_pct:.2f}%</span>'
        f'<span style="color:var(--amber);">Upper PI {upper_pct:.2f}%</span>'
        f'<span style="color:var(--red);">Effective {effective_pct:.2f}%</span>'
        f'<span style="color:var(--text-dim);">VIX {vix_pct:.2f}% · {confidence}% PI</span>'
        '</div>'
        '</div>'
    )


def _render_sf_strike_map_tier(
    tier: SpreadTier, plan: SpreadPlan, spx_ref: float,
    gex_ctx: GEXContext, selected_width: float = 25, ticker: str = "SPX",
    weekly_em: dict = None,
):
    """Swim-lane strike map grouped by semantic role.

    The old layout sorted every level (walls, shorts, spot, EM bounds,
    etc.) onto its own Y-row — clean for avoiding label overlap but
    forced the reader to mentally re-group "which of these is my trade
    vs structural GEX vs a forecast marker?" every time they scanned.

    This version keeps one shared price axis on X but splits the Y into
    three labeled lanes:

      TRADE   — Put Long / Put Short / Ref / Spot / Call Short / Call Long
      RANGE   — Effective bounds + EM bounds
      GEX     — Put Wall / Zero-Gamma / Call Wall

    Background shading:
      - tier-colored band between Put Short and Call Short (the "safe"
        window for the selected risk tier — where the spread profits)
      - translucent blue band across the weekly EM envelope

    Label collisions within a lane are avoided by alternating marker
    text positions top/bottom after sorting each lane's markers by
    price.
    """
    tier_colors = {"aggressive": "var(--red)", "moderate": "var(--amber)",
                   "conservative": "var(--green)"}
    tier_color = tier_colors.get(tier.risk_level, "var(--text-muted)")

    # Get strikes from the selected tier. We show the model's ORIGINAL strikes
    # (pre weekly-EM floor) — see the render tab for rationale. `.model_*` is
    # populated only when the floor moved the strike; fall back to the
    # (unchanged) floored value otherwise so the map matches the tables.
    call_short = tier.model_call_short if tier.model_call_short is not None else tier.call_short
    put_short  = tier.model_put_short  if tier.model_put_short  is not None else tier.put_short
    call_spreads = tier.model_call_spreads or tier.call_spreads
    put_spreads  = tier.model_put_spreads  or tier.put_spreads

    # Default long strikes from first spread
    call_long = call_spreads[0].long_strike if call_spreads else call_short + 25
    put_long  = put_spreads[0].long_strike  if put_spreads  else put_short - 25

    # Draw the same spread the table stars (best qualifying per side);
    # fall back to the model-recommended width when nothing qualifies.
    _best_c = _best_spread_idx(call_spreads)
    if _best_c is not None:
        call_long = call_spreads[_best_c].long_strike
    else:
        for s in call_spreads:
            if s.wing_width == selected_width:
                call_long = s.long_strike
    _best_p = _best_spread_idx(put_spreads)
    if _best_p is not None:
        put_long = put_spreads[_best_p].long_strike
    else:
        for s in put_spreads:
            if s.wing_width == selected_width:
                put_long = s.long_strike

    # Weekly expected move from Friday straddle (already computed by GEX engine)
    em_upper = (weekly_em or {}).get("upper_level", 0)
    em_lower = (weekly_em or {}).get("lower_level", 0)
    has_em = em_upper > 0 and em_lower > 0
    em_color = "#29b6f6"  # light blue

    # ── Swim-lane geometry (price → % position on one shared X axis) ──
    trade = [
        (put_long,      "Put Long",      "var(--red)",   "◄"),
        (put_short,     "Put Short",     tier_color,     "◆"),
        (spx_ref,       f"{ticker} Ref", "#ffffff",      "★"),
        (gex_ctx.spot,  "Spot",          "var(--amber)", "●"),
        (call_short,    "Call Short",    tier_color,     "◆"),
        (call_long,     "Call Long",     "var(--red)",   "►"),
    ]
    rng = [
        (plan.effective_lower_px, "Eff Lo", "var(--amber)", "▲"),
        (plan.effective_upper_px, "Eff Hi", "var(--amber)", "▲"),
    ]
    if has_em:
        rng += [(em_lower, "EM Lo", "var(--blue)", "▲"),
                (em_upper, "EM Hi", "var(--blue)", "▲")]
    gex = [
        (gex_ctx.put_wall,   "Put Wall",  "var(--red)",   "■"),
        (gex_ctx.zero_gamma, "Zero-Γ",    "var(--cyan)",  "✕"),
        (gex_ctx.call_wall,  "Call Wall", "var(--green)", "■"),
    ]

    all_px = [m[0] for m in trade + rng + gex]
    pmin, pmax = min(all_px), max(all_px)
    margin = (pmax - pmin) * 0.12 or 5.0
    axis_min, axis_max = pmin - margin, pmax + margin
    span = (axis_max - axis_min) or 1.0

    def _xp(px: float) -> float:
        return max(0.0, min(100.0, (px - axis_min) / span * 100.0))

    ticks = [round((axis_min + span * f) / 5) * 5 for f in (0, 0.25, 0.5, 0.75, 1.0)]

    def _lane(markers, fmt=lambda px: f"{px:,.0f}") -> str:
        out = []
        for i, (px, _lbl, color, glyph) in enumerate(sorted(markers, key=lambda m: m[0])):
            lbl_pos = "bottom:15px;" if i % 2 == 0 else "top:15px;"
            out.append(
                f'<div class="mk" style="left:{_xp(px):.2f}%;color:{color};">{glyph}'
                f'<span class="lbl" style="position:absolute;left:50%;transform:translateX(-50%);'
                f'{lbl_pos}color:{color};">{_lbl} {fmt(px)}</span></div>'
            )
        return "".join(out)

    grid = "".join(f'<div class="grid" style="left:{_xp(t):.2f}%;"></div>' for t in ticks)
    refs = (
        f'<div class="grid" style="left:{_xp(spx_ref):.2f}%;border-left:2px solid rgba(231,237,245,.4);"></div>'
        f'<div class="grid" style="left:{_xp(gex_ctx.spot):.2f}%;border-left:1.5px dashed var(--amber);"></div>'
    )
    safe = (
        f'<div class="safe" style="left:{_xp(put_short):.2f}%;'
        f'width:{max(0.0, _xp(call_short) - _xp(put_short)):.2f}%;'
        f'background:color-mix(in srgb,{tier_color} 12%,transparent);'
        f'border-left:1px solid {tier_color};border-right:1px solid {tier_color};"></div>'
    )
    em_band = ""
    if has_em:
        em_band = (
            f'<div class="safe" style="left:{_xp(em_lower):.2f}%;'
            f'width:{max(0.0, _xp(em_upper) - _xp(em_lower)):.2f}%;'
            'background:rgba(110,168,255,.06);border-left:1px dashed var(--blue);'
            'border-right:1px dashed var(--blue);"></div>'
        )
    xaxis = "".join(
        f'<span style="position:absolute;left:{_xp(t):.2f}%;transform:translateX(-50%);">{t:,.0f}</span>'
        for t in ticks
    )

    st.html(
        '<div class="sf-map"><div class="yax">'
        '<div class="lane">TRADE</div><div class="lane">RANGE</div><div class="lane">GEX</div>'
        '</div><div class="area">'
        f'{safe}{em_band}{grid}{refs}'
        f'<div class="lane-row">{_lane(trade, _fmt_strike)}</div>'
        f'<div class="lane-row">{_lane(rng)}</div>'
        f'<div class="lane-row">{_lane(gex)}</div>'
        '</div></div>'
        f'<div class="sf-map-x" style="position:relative;height:14px;">{xaxis}</div>'
        f'<div class="sf-map-foot">{ticker} PRICE LEVEL</div>'
    )


def _best_spread_idx(spreads) -> "int | None":
    """Index of the 'best' spread in a width ladder, or None.

    With the short strike fixed across the ladder, credit RATIO is always
    highest on the narrowest width (the marginal credit per extra point of
    wing decays), so 'max ratio' would just re-create an arbitrary
    always-first-row highlight. Instead: the WIDEST spread whose credit
    still clears MIN_CREDIT_RATIO and the event minimum width — i.e. the
    most absolute premium and breach headroom you can take while still
    being paid at least the floor per point of risk. In high-IV tape more
    rungs qualify and the highlight walks wider; on dead/thin chains
    (0.00 credits) nothing qualifies and nothing is highlighted.
    """
    best = None
    for i, s in enumerate(spreads):
        if not getattr(s, "meets_min_credit", False):
            continue
        if getattr(s, "below_min_width", False):
            continue
        if best is None or s.wing_width > spreads[best].wing_width:
            best = i
    return best


def _fmt_strike(x) -> str:
    """Format a strike / wing width / level, preserving genuine fractional
    strikes (e.g. NVDA's 207.5 on the $2.5 grid) instead of rounding them to a
    misleading integer, while keeping whole values clean (208, not 208.0).

    Without this the table formatted strikes with ``:,.0f``, rounding 207.5 → 208
    so a true 2.5-wide spread looked 2-wide next to its "W = 2.5pt" label."""
    if x is None:
        return "—"
    x = float(x)
    if x == round(x):
        return f"{int(round(x)):,}"
    return f"{x:,.2f}".rstrip("0").rstrip(".")


def _render_sf_spread_table(spreads):
    """Render the spread ladder as an 8-column terminal table, highlighting the
    best qualifying spread (see _best_spread_idx). Presentation only — every
    value comes straight off the (untouched) SpreadOption objects."""
    if not spreads:
        st.html('<div class="sf-note">No spreads available.</div>')
        return

    best_idx = _best_spread_idx(spreads)

    body = []
    for i, s in enumerate(spreads):
        wlabel = f"{_fmt_strike(s.wing_width)}pt"
        if getattr(s, "below_min_width", False):
            wlabel += "*"
        if i == best_idx:
            wlabel = "★ " + wlabel
        cr = f"{s.estimated_credit:.2f}" + ("*" if getattr(s, "credit_source", "bsm") == "bsm" else "")
        ok = "Y" if s.meets_min_credit else "N"
        ok_cls = "oky" if s.meets_min_credit else "okn"
        body.append(
            f'<tr class="{"best" if i == best_idx else ""}">'
            f'<td>{wlabel}</td>'
            f'<td class="short">{_fmt_strike(s.short_strike)}</td>'
            f'<td class="long">{_fmt_strike(s.long_strike)}</td>'
            f'<td class="cr">{cr}</td>'
            f'<td>${s.max_loss:,.0f}</td>'
            f'<td>{s.breakeven:,.0f}</td>'
            f'<td>{s.credit_ratio:.1%}</td>'
            f'<td class="{ok_cls}">{ok}</td>'
            '</tr>'
        )

    st.html(
        '<div class="sf-table-wrap"><table class="sf-table"><thead><tr>'
        '<th>W</th><th>Short</th><th>Long</th><th>CR</th><th>MaxLoss</th>'
        '<th>BE</th><th>R%</th><th>OK</th>'
        '</tr></thead><tbody>' + "".join(body) + '</tbody></table></div>'
    )
    if best_idx is not None:
        st.caption(
            f"★ Best = widest wing still paying ≥{MIN_CREDIT_RATIO:.0%} of width "
            "(max premium + breach headroom at an acceptable credit-per-risk floor)."
        )
    else:
        st.caption(
            f"No spread clears the {MIN_CREDIT_RATIO:.0%} credit-to-width floor — "
            "credit is too thin at these strikes/widths to be worth the risk."
        )
