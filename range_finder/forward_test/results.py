"""Read models, denominator-aware statistics and the application's Excel export."""
from collections import defaultdict
from io import BytesIO
import json

from range_finder.recommendations import TIER_KEYS, TIER_LABELS


def load_results(store, study_id):
    rows = store.query("""
        SELECT s.*, f.forecast_id, f.tier, f.payload_json AS forecast_json,
               c.payload_json AS score_json, c.scored_at, c.revision AS score_revision
        FROM ft_slots s LEFT JOIN ft_forecasts f ON f.slot_id=s.slot_id
        LEFT JOIN ft_scores c ON c.forecast_id=f.forecast_id AND c.revision=(
            SELECT MAX(c2.revision) FROM ft_scores c2 WHERE c2.forecast_id=f.forecast_id)
        WHERE s.study_id=? ORDER BY s.week_start DESC, s.ticker, s.model, f.tier
    """, (study_id,))
    result = []
    for row in rows:
        base = {k: row[k] for k in ("study_id", "week_start", "ticker", "model", "model_version", "cohort",
                                    "attempts", "error", "forecast_id", "scored_at", "score_revision")}
        if row["forecast_json"]:
            f = json.loads(row["forecast_json"])
            score = json.loads(row["score_json"]) if row["score_json"] else {}
            result.append({**base, **f, **score, "capture_status": row["status"],
                           "status": score.get("status", "pending"),
                           "classification": score.get("classification", "Pending")})
        else:
            for tier, label in zip(TIER_KEYS, TIER_LABELS):
                result.append({**base, "tier": tier, "tier_label": label, "status": row["status"],
                               "capture_status": row["status"], "classification": row["status"].capitalize(),
                               "close_eligible": False, "path_eligible": False})
    return result


def metrics(rows):
    close = [r for r in rows if r.get("close_eligible") is True]
    path = [r for r in rows if r.get("path_eligible") is True]
    breached_close = [r for r in path if r.get("close_eligible") and r.get("either_breach")]
    widths = [r["range_width_ratio"] for r in rows if r.get("range_width_ratio") is not None]
    hits = sum(r.get("close_inside") is True for r in close)
    breaches = sum(r.get("either_breach") is True for r in path)
    recovered = sum(r.get("returned_inside") is True for r in breached_close)
    return {"close_n": len(close), "close_hits": hits, "close_hit_rate": hits / len(close) if close else None,
            "path_n": len(path), "breaches": breaches, "breach_rate": breaches / len(path) if path else None,
            "recovery_n": len(breached_close), "recoveries": recovered,
            "recovery_rate": recovered / len(breached_close) if breached_close else None,
            "put_failures": sum(r["final_close"] < r["put_short"] for r in close),
            "call_failures": sum(r["final_close"] > r["call_short"] for r in close),
            "pending": sum(r.get("status") == "pending" for r in rows),
            "missed": sum(r.get("status") == "missed" for r in rows),
            "unavailable": sum(r.get("status") == "unavailable" for r in rows),
            "incomplete": sum(r.get("status") in ("incomplete", "invalid") for r in rows),
            "average_width_ratio": sum(widths) / len(widths) if widths else None, "width_n": len(widths)}


GROUP_KEYS = ("study_id", "cohort", "model_version", "ticker", "model", "tier")


def scoreboard(rows):
    groups = defaultdict(list)
    for row in rows:
        groups[tuple(row.get(k) for k in GROUP_KEYS)].append(row)
    ordered = sorted(groups.items(), key=lambda item: (item[0][3], item[0][4],
        TIER_KEYS.index(item[0][5]), item[0][0], item[0][1], item[0][2]))
    return [{**dict(zip(GROUP_KEYS, key)), **metrics(values)} for key, values in ordered]


def build_workbook(rows):
    """Runtime export of durable results, following the supplied weekly masters.

    Reuses the application's existing openpyxl dependency. No import of manual
    workbooks, live formulas, implied P&L or user-entered outcome cells.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    subtitle = ("Deterministic fixture data. Synthetic prices and outcomes; excluded from the prospective study."
                if rows and all(r.get('cohort') == 'deterministic_fixture' for r in rows)
                else "Durable observational records. Close inside includes equality. No trade returns are calculated.")
    labels = dict(zip(TIER_KEYS, TIER_LABELS))
    versions = {r.get('model_version', '') for r in rows}
    # Keep the visible version label short without merging distinct identities.
    prefix = 12
    while len({v[:prefix] for v in versions}) != len(versions):
        prefix += 1

    def safe(value):
        if isinstance(value, (dict, list)):
            return json.dumps(value, sort_keys=True)
        if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
            return "'" + value
        return value

    def table(name, title, records, keys, labels=None):
        ws = wb.create_sheet(name)
        ws.append([title])
        ws.append([subtitle])
        ws.append([])
        ws.append(labels or [k.replace("_", " ").capitalize() for k in keys])
        for record in records:
            ws.append([safe(record.get(k)) for k in keys])
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "D5"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, min(8, len(keys))))
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(2, min(8, len(keys))))
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 26
        ws['A2'].font = Font(name="Arial", size=10, color="526579")
        ws['A2'].alignment = Alignment(wrap_text=True, vertical="center")
        ws.auto_filter.ref = f"A4:{get_column_letter(len(keys))}{max(ws.max_row, 4)}"
        for c in ws[1]:
            c.font = Font(name="Arial", size=14, bold=True)
        for c in ws[4]:
            c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="24435B")
            c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[4].height = 42
        for n, key in enumerate(keys, 1):
            ws.column_dimensions[get_column_letter(n)].width = 18 if len(key) > 14 else 15
            if key in ('model', 'tier_label', 'cohort'):
                ws.column_dimensions[get_column_letter(n)].width = 27
            if key in ('model_version', 'forecast_id'):
                ws.column_dimensions[get_column_letter(n)].width = 74
            if key.endswith('_at') or key.endswith('_cutoff'):
                ws.column_dimensions[get_column_letter(n)].width = 31
            if key in ("classification", "error", "path_limitations", "interpretation"):
                ws.column_dimensions[get_column_letter(n)].width = 48
            for cells in ws.iter_rows(min_row=5, min_col=n, max_col=n):
                c = cells[0]
                c.font = Font(name="Arial", size=10)
                c.alignment = Alignment(vertical='center')
                if c.row % 2:
                    c.fill = PatternFill('solid', fgColor='F0F5F8')
                if key in ('close_inside', 'lower_pi_inside', 'point_inside', 'pi_upper_inside', 'effective_inside') and isinstance(c.value, bool):
                    c.fill = PatternFill('solid', fgColor='DDECE5' if c.value else 'FBE8D4')
                if isinstance(c.value, (float, int)) and not isinstance(c.value, bool):
                    c.number_format = "0.0%" if key.endswith(("_rate", "_ratio")) else "#,##0.00" if isinstance(c.value, float) else "#,##0"
        for row in range(5, ws.max_row+1):
            ws.row_dimensions[row].height = 22
        ws.print_title_rows = '1:4'
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        return ws

    score = scoreboard(rows)
    for row in score:
        row['tier_label'] = labels[row['tier']]
        row['version_label'] = row['model_version'][:prefix]
    score_keys = ['ticker', 'model', 'tier_label', 'close_hit_rate', 'close_n', 'breach_rate', 'path_n',
                  'recovery_rate', 'recovery_n', 'average_width_ratio', 'width_n', 'put_failures', 'call_failures',
                  'pending', 'missed', 'unavailable', 'incomplete', 'version_label', 'cohort', 'study_id',
                  'model_version', 'close_hits', 'breaches', 'recoveries']
    table("Scoreboard", "Weekly forward test scoreboard", score, score_keys)
    weekly_keys = ["ticker", "model", "reference", "final_close",
                   "lower_pi_put", "lower_pi_call", "point_put", "point_call",
                   "pi_upper_put", "pi_upper_call", "effective_put", "effective_call",
                   "lower_pi_inside", "point_inside", "pi_upper_inside", "effective_inside", "expiration", "capture_status",
                   "version_label", "cohort", "model_version"]
    for week in sorted({r["week_start"] for r in rows}, reverse=True):
        groups = defaultdict(dict)
        for row in (r for r in rows if r["week_start"] == week):
            key = tuple(row.get(k) for k in GROUP_KEYS[:-1])
            merged = groups[key]
            for k in ('ticker', 'model', 'reference', 'final_close', 'expiration', 'capture_status', 'cohort', 'model_version'):
                merged[k] = row.get(k)
            merged['version_label'] = row['model_version'][:prefix]
            for field, source in (("put", "put_short"), ("call", "call_short"), ("inside", "close_inside")):
                merged[f"{row['tier']}_{field}"] = row.get(source)
        table(week, f"Week of {week}", list(groups.values()), weekly_keys)
    detail_keys = ["week_start", "ticker", "model", "tier_label", "cohort", "model_version", "put_short", "call_short",
                   "final_close", "close_inside", "close_eligible", "path_eligible", "classification",
                   "put_touch", "call_touch", "put_breach", "call_breach", "either_breach", "both_breached",
                   "earlier_close_outside", "returned_inside", "close_on_boundary", "observed_low", "observed_high",
                   "whole_week_low", "whole_week_high", "whole_week_complete", "first_breach_session",
                   "first_breach_timestamp", "first_breach_interval", "first_breach_is_earliest_verified",
                   "window_start", "window_end", "scheduled_at", "captured_at", "available_at", "feature_cutoff",
                   "training_cutoff", "data_delay_seconds", "expiration", "contract_root", "settlement_convention",
                   "settlement_status", "settlement_value", "settlement_inside", "interpretation", "sources",
                   "path_limitations", "scored_at", "score_revision", "forecast_id", "error", "wings"]
    table("Breach details", "Weekly close and prospective breach evidence", rows, detail_keys)
    notes = [
        ("Close sample", "Valid final regular-session daily OHLC; inclusive put <= close <= call. Independent of path completeness."),
        ("Breach sample", "Complete post-availability coverage and known strict-breach flags on both sides. Missing intervals do not mean no breach."),
        ("Recovery sample", "Rows with valid close and path data and a strict breach. Recovery rate = inside final close / these breached rows."),
        ("Touches", "Low <= put or high >= call. Equality alone is not a strict breach. Partial-minute equality may be unknown."),
        ("First breach", "Session and interval only where supported. Exact timestamps remain blank for OHLC evidence."),
        ("Extremes", "Observed extremes use verified post-availability intervals. Whole-week daily extremes include pre-forecast movement."),
        ("Versions", "Different cohorts and model versions remain separate in the Scoreboard. Counts are tier records."),
        ("Version labels", "Short version labels identify groups. Full model version hashes are retained in each data sheet."),
        ("Settlement", "Official contract settlement is separate from the weekly close. Stock/ETF results are theoretical expiration proxies."),
        ("Source", "https://docs.tradier.com/reference/brokerage-api-markets-get-timesales"),
        ("SPX contracts", "https://www.cboe.com/available_weeklys/"),
        ("Legacy masters", "M2_Vix_Master.xlsx and M3_Extended_Master.xlsx are layout references only; none of their rows were imported."),
    ]
    table("Read me", "Evaluation definitions", [{"topic": k, "definition": v} for k, v in notes], ["topic", "definition"])
    wb["Read me"].column_dimensions["B"].width = 115
    for row in wb["Read me"].iter_rows(min_row=5):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        wb["Read me"].row_dimensions[row[0].row].height = 32
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
