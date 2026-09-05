from concurrent.futures import ThreadPoolExecutor
from copy import deepcopy
from datetime import date, timedelta
from io import BytesIO
import json
import sqlite3

import numpy as np
import pandas as pd
import pytest

from range_finder.forward_test.capture import capture_model, validate_saved_fit
from range_finder.forward_test.config import MODELS, UNIVERSE
from range_finder.forward_test.store import Store
from range_finder.forward_test.runner import run_study
from range_finder.forward_test.scoring import score_forecast
from range_finder.forward_test.results import load_results, metrics, scoreboard, build_workbook
from range_finder.recommendations import TIER_KEYS, build_recommendations, tier_bands, chain_entry_to_quotes
from range_finder.trading_week import trading_week, listed_week_expiration
from range_finder.tests.forward_fixtures import Clock, FixtureProvider, prepared_fixture, scoring_fixture


@pytest.fixture
def week():
    return trading_week(date(2026, 9, 7))  # Labor Day -> Tuesday


@pytest.fixture
def store(tmp_path, week):
    s = Store(sqlite3.connect(tmp_path / 'isolated.sqlite', isolation_level=None))
    s.migrate()
    s.register('fixture', str(week.monday), week.capture_start, {'fixture': True})
    yield s
    s.close()


@pytest.mark.parametrize('ticker', UNIVERSE)
@pytest.mark.parametrize('model', MODELS)
def test_all_models_tiers_ui_parity_no_gex(ticker, model, week):
    clock = Clock(week.capture_start + timedelta(seconds=17))
    p = prepared_fixture(ticker, week, clock)
    inputs, frozen = capture_model(p, model, week, 'v1', clock())
    from range_finder.har_model import fit_production_model
    from ui_spread_finder import _tier_bands_from_tiers
    training = p['features'][p['features'].index < pd.Timestamp(week.monday)]
    cols = inputs['feature_columns']
    fit = fit_production_model(training, cols)
    _, _, tiers = build_recommendations(result=fit, feature_row=p['features'].loc[pd.Timestamp(week.monday)],
                                       feature_cols=cols, reference=p['reference'], vix=p['vix'], ticker=ticker,
                                       week_start=str(week.monday), side_share_q=inputs['side_share']['q'],
                                       chain_quotes=chain_entry_to_quotes(p['chain']))
    assert _tier_bands_from_tiers(tiers) == {r['tier']: (r['put_short'],r['call_short']) for r in frozen}
    assert [f['tier'] for f in frozen] == list(TIER_KEYS)
    assert all(f['wings']['put'] and f['wings']['call'] for f in frozen)
    p['features']['gex_normalized'] = -1e10
    p['features']['gex_flag'] = -1
    _, other = capture_model(p, model, week, 'v1', clock())
    assert [(r['put_short'], r['call_short'], r['buffer_pct']) for r in frozen] == [(r['put_short'], r['call_short'], r['buffer_pct']) for r in other]


def test_no_lookahead_and_no_retrospective_admission(week):
    clock = Clock(week.capture_start)
    p = prepared_fixture('SPY', week, clock)
    before, f1 = capture_model(p, 'M3_extended', week, 'v1', clock())
    p['features'].loc[pd.Timestamp(week.monday), ['log_range','range_pct']] = [999.,999.]
    p['features'].loc[pd.Timestamp(week.monday)+pd.Timedelta(days=7)] = p['features'].iloc[-2]
    after, f2 = capture_model(p, 'M3_extended', week, 'v1', clock())
    assert before['fit'] == after['fit']
    assert [(r['put_short'],r['call_short']) for r in f1] == [(r['put_short'],r['call_short']) for r in f2]
    with pytest.raises(ValueError, match='Missed'):
        capture_model(p, 'M3_extended', week, 'v1', week.capture_end)


def test_saved_fit_rejects_unknown_identity_and_gex(week):
    from range_finder.model_persistence import IncompatibleModelError, SCHEMA_VERSION
    base = {'schema_version': SCHEMA_VERSION, 'ticker':'SPY','model_name':'M2_vix',
            'training_cutoff':str(week.monday-timedelta(days=7)), 'feature_cols':['har_d1']}
    validate_saved_fit(base, 'SPY','M2_vix',week)
    for field, value in [('schema_version', 0), ('ticker','SPX'), ('feature_cols',['gex_normalized']), ('training_cutoff',str(week.monday))]:
        with pytest.raises(IncompatibleModelError):
            validate_saved_fit({**base,field:value}, 'SPY','M2_vix',week)


@pytest.mark.parametrize('d,first,last,hours', [
    ('2026-09-07','2026-09-08','2026-09-11',6.5),
    ('2026-03-30','2026-03-30','2026-04-02',6.5),
    ('2026-11-23','2026-11-23','2026-11-27',3.5),
    ('2026-03-09','2026-03-09','2026-03-13',6.5),
    ('2026-11-02','2026-11-02','2026-11-06',6.5)])
def test_calendar_expiration_dst(d,first,last,hours):
    week=trading_week(date.fromisoformat(d))
    assert str(week.sessions[0].day)==first
    assert str(week.sessions[-1].day)==last
    assert (week.sessions[-1].close-week.sessions[-1].open).total_seconds()/3600==hours
    assert listed_week_expiration([last],week)==last
    assert listed_week_expiration([str(week.monday+timedelta(days=7))],week) is None
    assert week.capture_start.astimezone(__import__('zoneinfo').ZoneInfo('America/New_York')).strftime('%H:%M')=='09:45'


@pytest.mark.parametrize('put_breach,call_breach',[(False,False),(True,False),(False,True),(True,True)])
def test_inside_and_each_recovery(week,put_breach,call_breach):
    f, obs, clock=scoring_fixture(week)
    day=str(week.sessions[1].day)
    obs[day]['daily']['low']=94. if put_breach else 96.
    obs[day]['daily']['high']=106. if call_breach else 104.
    s=score_forecast(f,week,obs,clock())
    assert s['close_inside'] and s['path_eligible']
    assert s['put_breach']==put_breach and s['call_breach']==call_breach
    assert s['returned_inside']==(put_breach or call_breach)
    assert s['first_breach_timestamp'] is None


@pytest.mark.parametrize('close,inside,boundary',[(94,False,False),(106,False,False),(95,True,True),(105,True,True)])
def test_close_failures_and_equality(week,close,inside,boundary):
    f,obs,clock=scoring_fixture(week)
    final=obs[str(week.sessions[-1].day)]['daily']
    final.update(close=close,low=min(95,close),high=max(105,close))
    s=score_forecast(f,week,obs,clock())
    assert s['close_inside']==inside and s['close_on_boundary']==boundary
    if inside:
        assert s['put_touch'] and s['call_touch']
        assert not s['either_breach']


def test_earlier_outside_close_recovers(week):
    f,obs,clock=scoring_fixture(week)
    obs[str(week.sessions[1].day)]['daily'].update(close=94.,low=93.)
    s=score_forecast(f,week,obs,clock())
    assert s['close_inside'] and s['earlier_close_outside'] and s['returned_inside']


def test_close_eligible_with_missing_path(week):
    f,obs,clock=scoring_fixture(week)
    del obs[str(week.sessions[1].day)]
    s=score_forecast(f,week,obs,clock())
    assert s['close_eligible'] and not s['path_eligible']
    assert s['put_breach'] is None and s['call_breach'] is None
    m=metrics([{**f,**s}])
    assert m['close_n']==1 and m['path_n']==0 and m['breach_rate'] is None


def test_partial_minute_cannot_assign_preforecast_breach(week):
    f,obs,clock=scoring_fixture(week,available=week.capture_start+timedelta(seconds=20))
    first=obs[str(week.sessions[0].day)]
    first['daily']['low']=90.
    first['minutes'][0]['low']=90.
    s=score_forecast(f,week,obs,clock())
    assert s['put_breach'] is None and not s['path_eligible']
    assert s['whole_week_low']==90. and s['observed_low']==96.
    first['minutes'][2]['low']=94.
    s=score_forecast(f,week,obs,clock())
    assert s['put_breach'] is True and s['first_breach_session']==str(week.sessions[0].day)


def test_partial_minute_last_trade_may_precede_availability(week):
    f,obs,clock=scoring_fixture(week,available=week.capture_start+timedelta(seconds=20))
    first=obs[str(week.sessions[0].day)]
    first['daily']['low']=90.
    first['minutes'][0].update(low=90.,close=94.)
    result=score_forecast(f,week,obs,clock())
    assert result['put_breach'] is None and result['first_breach_session'] is None
    assert not result['path_eligible'] and result['close_eligible']


def test_naive_minute_and_early_prior_session_do_not_establish_absence(week):
    f,obs,clock=scoring_fixture(week)
    first=obs[str(week.sessions[0].day)]
    first['minutes'][0]['start']=first['minutes'][0]['start'].split('+')[0]
    result=score_forecast(f,week,obs,clock())
    assert result['close_eligible'] and not result['path_eligible']
    first['collected_at']=(week.sessions[0].close-timedelta(minutes=1)).isoformat()
    assert score_forecast(f,week,obs,clock())['earlier_close_outside'] is None


def test_conflicting_duplicate_minute_is_not_breach_evidence(week):
    f,obs,clock=scoring_fixture(week)
    first=obs[str(week.sessions[0].day)]
    first['minutes'].append({**first['minutes'][1],'low':90.})
    result=score_forecast(f,week,obs,clock())
    assert not result['path_eligible'] and result['put_breach'] is None


def test_no_whole_day_preforecast_breach_claim(week):
    f,obs,clock=scoring_fixture(week)
    obs[str(week.sessions[0].day)]['daily']['low']=80.
    s=score_forecast(f,week,obs,clock())
    assert s['put_breach'] is False and s['whole_week_low']==80. and s['observed_low']==96.


def test_stale_bars_invalid_identity_and_settlement(week):
    f,obs,clock=scoring_fixture(week)
    final=obs[str(week.sessions[-1].day)]
    s=score_forecast(f,week,obs,clock())
    assert s['settlement_status']=='unverified' and s['settlement_value'] is None
    final['settlement']={'status':'verified','value':99.,'root':'SPY','expiration':f['expiration'],'source':'official fixture'}
    assert score_forecast(f,week,obs,clock())['settlement_inside'] is True
    final['daily']['date']=str(week.sessions[-2].day)
    assert not score_forecast(f,week,obs,clock())['close_eligible']
    final['ticker']='SPX'
    assert not score_forecast(f,week,obs,clock())['close_eligible']


def test_end_to_end_and_immutable_duplicate_trigger(store,week):
    clock=Clock(week.capture_start+timedelta(seconds=20))
    provider=FixtureProvider(clock)
    result=run_study(store,provider,'fixture',clock=clock,model_version='fixture-v1')
    assert not result['errors'] and result['captured_models']==16
    frozen=store.forecasts('fixture')
    assert len(frozen)==64
    provider.prepares.clear()
    assert run_study(store,provider,'fixture',clock=clock,model_version='new-version')['captured_models']==0
    assert not provider.prepares and store.forecasts('fixture')==frozen
    with pytest.raises(sqlite3.IntegrityError,match='immutable'):
        with store.transaction():
            store.execute("UPDATE ft_forecasts SET payload_json='{}'")
    clock.value=week.evaluation_close+timedelta(hours=3)
    run_study(store,provider,'fixture',clock=clock,model_version='fixture-v1')
    records=load_results(store,'fixture')
    assert len(records)==64 and metrics(records)['close_n']==64 and metrics(records)['path_n']==64
    assert len(provider.observes)==4*len(week.sessions)
    workbook=build_workbook(records)
    from openpyxl import load_workbook
    wb=load_workbook(BytesIO(workbook),data_only=False)
    assert wb.sheetnames==['Scoreboard',str(week.monday),'Breach details','Read me']
    assert wb['Breach details'].max_row==68
    assert sum(r['close_n'] for r in scoreboard(records))==64
    assert store.forecasts('fixture')==frozen


def test_partial_failure_safe_retry_and_missed_window(store,week):
    clock=Clock(week.capture_start)
    provider=FixtureProvider(clock)
    provider.fail_tickers={'AMD'}
    run_study(store,provider,'fixture',clock=clock,model_version='v1')
    assert len(store.forecasts('fixture'))==48
    provider.fail_tickers.clear()
    run_study(store,provider,'fixture',clock=clock,model_version='v1')
    assert len(store.forecasts('fixture'))==64
    next_week=trading_week(week.monday+timedelta(days=7))
    clock.value=next_week.capture_end
    before=len(provider.prepares)
    run_study(store,provider,'fixture',clock=clock,model_version='v1')
    assert all(s['status']=='missed' for s in store.slots('fixture',str(next_week.monday)))
    assert len(provider.prepares)==before


def test_concurrent_capture_one_winner(tmp_path,week):
    path=tmp_path/'concurrent.sqlite'
    root=Store(sqlite3.connect(path,isolation_level=None))
    root.migrate(); root.register('x',str(week.monday),week.capture_start,{'fixture':True})
    root.seed_week('x',str(week.monday),'v1',week.capture_start)
    slot=root.slots('x',str(week.monday))[0]
    p=prepared_fixture(slot['ticker'],week,Clock(week.capture_start))
    inputs,forecasts=capture_model(p,slot['model'],week,'v1',week.capture_start)
    def write(_):
        s=Store(sqlite3.connect(path,isolation_level=None,timeout=10))
        try:return s.freeze(slot,inputs,forecasts,week.capture_start,week)
        finally:s.close()
    with ThreadPoolExecutor(max_workers=4) as pool:
        assert sum(pool.map(write,range(4)))==1
    assert len(root.forecasts('x'))==4
    root.close()


def test_versioned_correction_can_return_to_prior_value(store,week):
    now=week.evaluation_close+timedelta(hours=3)
    a=store.append_observation('fixture',str(week.monday),'SPY',str(week.sessions[-1].day),{'value':1},now)
    same=store.append_observation('fixture',str(week.monday),'SPY',str(week.sessions[-1].day),{'value':1},now)
    assert a==same
    store.append_observation('fixture',str(week.monday),'SPY',str(week.sessions[-1].day),{'value':2},now)
    c=store.append_observation('fixture',str(week.monday),'SPY',str(week.sessions[-1].day),{'value':1},now)
    assert c!=a
    assert store.observations('fixture',str(week.monday),'SPY')[str(week.sessions[-1].day)]['value']==1


def test_retry_budget_and_delayed_observations(store,week):
    clock=Clock(week.capture_start); provider=FixtureProvider(clock)
    provider.fail_tickers=set(UNIVERSE)
    for _ in range(5):run_study(store,provider,'fixture',clock=clock,model_version='v1')
    assert len(provider.prepares)==12  # three attempts per ticker, never five
    assert all(s['attempts']==3 for s in store.slots('fixture',str(week.monday)))


def test_delayed_outcomes_reconcile_once_per_day_and_stop_after_deadline(store,week,monkeypatch):
    clock=Clock(week.capture_start); provider=FixtureProvider(clock)
    run_study(store,provider,'fixture',clock=clock,model_version='v1')
    original=store.forecasts('fixture')
    observe=provider.observe
    def delayed(ticker,session,first_available=None):
        result=observe(ticker,session,first_available)
        if session==week.sessions[-1]:result['daily']=None
        return result
    monkeypatch.setattr(provider,'observe',delayed)
    clock.value=week.evaluation_close+timedelta(hours=3)
    assert run_study(store,provider,'fixture',clock=clock,model_version='v1')['errors']
    assert metrics(load_results(store,'fixture'))['close_n']==0
    checked=len(provider.observes)
    run_study(store,provider,'fixture',clock=clock,model_version='v1')
    assert len(provider.observes)==checked
    monkeypatch.setattr(provider,'observe',observe)
    clock.value+=timedelta(days=1)
    run_study(store,provider,'fixture',clock=clock,model_version='v1')
    assert metrics(load_results(store,'fixture'))['close_n']==64
    assert store.forecasts('fixture')==original
    checked=len(provider.observes)
    clock.value=week.evaluation_close+timedelta(days=15)
    run_study(store,provider,'fixture',clock=clock,model_version='v1')
    assert len(provider.observes)==checked


def test_scorer_rejects_incomplete_minute_coverage_and_early_close_data(week):
    f,obs,clock=scoring_fixture(week)
    first=obs[str(week.sessions[0].day)]
    first['minutes'].pop(15)
    s=score_forecast(f,week,obs,clock())
    assert s['close_eligible'] and not s['path_eligible'] and s['put_breach'] is None
    final=obs[str(week.sessions[-1].day)]
    final['collected_at']=(week.evaluation_close+timedelta(minutes=20)).isoformat()
    assert not score_forecast(f,week,obs,clock())['close_eligible']


def test_correct_denominators_across_versions_and_excel_literals(week):
    f,obs,clock=scoring_fixture(week)
    complete=score_forecast(f,week,obs,clock())
    del obs[str(week.sessions[1].day)]
    incomplete=score_forecast(f,week,obs,clock())
    base={**f,'study_id':'x','cohort':'primary','model':'M2_vix','model_version':'v1','week_start':str(week.monday),
          'ticker':'SPY','tier':'point','range_width_ratio':.10,'tier_label':'Point Estimate'}
    rows=[{**base,**complete},{**base,**incomplete},{**base,'model_version':'v2','status':'missed','error':'=DO_NOT_EXECUTE()'}]
    m=metrics(rows)
    assert (m['close_n'],m['path_n'],m['missed'])==(2,1,1)
    assert len(scoreboard(rows))==2
    from openpyxl import load_workbook
    w=load_workbook(BytesIO(build_workbook(rows)))
    assert all(c.data_type!='f' for s in w for row in s for c in row)


def test_feature_subset_has_distinct_version(week):
    clock=Clock(week.capture_start); p=prepared_fixture('AMD',week,clock)
    _,full=capture_model(p,'M4_full',week,'methodology-v1',clock())
    p['features']['yield_spread']=np.nan
    _,subset=capture_model(p,'M4_full',week,'methodology-v1',clock())
    assert full[0]['model_version']!=subset[0]['model_version']


def test_two_week_scoreboard_accumulates_refits_but_separates_methodology_changes(store,week):
    from range_finder.feature_builder import build_features
    from range_finder.forward_test.provider import frame_records
    from range_finder.forward_test.config import methodology
    stable_id=methodology()[0]
    forecast_ids=[]
    for number in range(3):
        current=trading_week(week.monday+timedelta(weeks=number))
        clock=Clock(current.capture_start+timedelta(seconds=17))
        p=prepared_fixture('SPY',current,clock)
        # A different completed training outcome changes the actual regression,
        # not just a label or a mocked fit hash. Rebuild with production features.
        raw={k:pd.DataFrame(v).set_index('date') if v else pd.DataFrame() for k,v in p['raw_inputs'].items()}
        for frame in raw.values():
            if not frame.empty:frame.index=pd.to_datetime(frame.index)
        last=raw['weekly'].index[-1]
        width=float(raw['weekly'].loc[last,'range_pct'])+.007*number
        raw['weekly'].loc[last,['spx_high','range_pct','log_range']]=[
            raw['weekly'].loc[last,'spx_low']+600*width,width,np.log(width)]
        p['weekly']=raw['weekly']
        p['features']=build_features(None,ticker='SPY',inputs=raw,as_of=clock(),persist=False)
        p['raw_inputs']={k:frame_records(v) for k,v in raw.items()}
        # A real change in the actual M2 design matrix creates the third
        # comparison group. The test does not manufacture a different ID.
        if number==2:
            p['features']['vix_close']=np.nan
        rules=stable_id
        store.seed_week('fixture',str(current.monday),rules,clock())
        slot=next(s for s in store.slots('fixture',str(current.monday)) if s['ticker']=='SPY' and s['model']=='M2_vix')
        inputs,forecasts=capture_model(p,'M2_vix',current,rules,clock())
        assert store.freeze(slot,inputs,forecasts,clock(),current)
        clock.value=current.evaluation_close+timedelta(hours=3)
        provider=FixtureProvider(clock)
        for session in current.sessions:
            obs=provider.observe('SPY',session,current.capture_start+timedelta(seconds=17) if session==current.sessions[0] else None)
            store.append_observation('fixture',str(current.monday),'SPY',str(session.day),obs,clock())
        observations=store.observations('fixture',str(current.monday),'SPY')
        for f in store.forecasts('fixture',str(current.monday)):
            payload=json.loads(f['payload_json'])
            store.append_score(f['forecast_id'],score_forecast(payload,current,observations,clock()),'test',clock())
            if f['tier']=='point':forecast_ids.append(f)
    rows=[r for r in load_results(store,'fixture') if r.get('forecast_id') and r['tier']=='point']
    groups=scoreboard(rows)
    assert sorted(g['close_n'] for g in groups)==[1,2]
    assert sorted(g['path_n'] for g in groups)==[1,2]
    payloads=[json.loads(f['payload_json']) for f in forecast_ids]
    assert payloads[0]['model_version']==payloads[1]['model_version']!=payloads[2]['model_version']
    assert payloads[0]['fit_id']!=payloads[1]['fit_id']
    snapshots=[json.loads(store.query('SELECT payload_json FROM ft_inputs WHERE input_id=?',(f['input_id'],))[0]['payload_json']) for f in forecast_ids]
    assert snapshots[0]['fit']['parameters']!=snapshots[1]['fit']['parameters']
    assert snapshots[0]['fit']['covariance']!=snapshots[1]['fit']['covariance']
    assert snapshots[0]['feature_columns']==snapshots[1]['feature_columns']!=snapshots[2]['feature_columns']
    assert snapshots[0]['features']!=snapshots[1]['features']
    assert snapshots[0]['training_cutoff']!=snapshots[1]['training_cutoff']
    assert snapshots[0]['source_inputs_id']!=snapshots[1]['source_inputs_id']
